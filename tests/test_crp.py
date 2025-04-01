import unittest
import os
import pandas as pd
from openpyxl import Workbook
from crp import CRP


class TestCRPInicializacao(unittest.TestCase):
    """
    Classe de teste para verificar a inicialização do CRP.
    """

    def setUp(self):
        """
        Método que é executado antes de cada teste.
        Configura o ambiente de teste, criando uma instância da classe CRP.
        """
        self.pasta_testes = "test_data_crp"

        # Cria o diretório de testes se não existir
        if not os.path.exists(self.pasta_testes):
            os.makedirs(self.pasta_testes)

        # Cria os arquivos necessários para o teste
        self.criar_arquivo_planejamento_teste()

        # Inicializa o CRP com o arquivo de planejamento
        self.crp = CRP(self.pasta_testes)

    def tearDown(self):
        """
        Método que é executado após cada teste.
        Limpa o ambiente de teste, removendo os arquivos e diretórios criados.
        """
        # Remove o diretório de testes e seu conteúdo
        for arquivo in os.listdir(self.pasta_testes):
            caminho_arquivo = os.path.join(self.pasta_testes, arquivo)
            os.remove(caminho_arquivo)
        os.rmdir(self.pasta_testes)

    def criar_arquivo_planejamento_teste(self):
        """
        Cria um arquivo de planejamento de teste para ser usado nos testes.
        """
        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active

        # Adicionar cabeçalhos
        ws.append(["Material", "Estoque Atual", "2023-10-01", "2023-10-15", "2023-10-30"])

        # Adicionar dados
        ws.append(["ETF", 5, 100, 50, 75])
        ws.append(["ETI", 10, 80, 60, 90])
        ws.append(["JOKER", 15, 200, 100, 150])
        ws.append(["DAQ", 20, 150, 75, 125])

        # Salvar o arquivo
        wb.save(os.path.join(self.pasta_testes, "planejamento_mrp.xlsx"))

    def test_inicializar_crp(self):
        """
        Testa se o CRP é inicializado corretamente.
        """
        # Verifica se o CRP foi inicializado
        self.assertIsNotNone(self.crp)

        # Verifica se o diretório de arquivos foi definido corretamente
        self.assertEqual(self.crp.pasta_arquivos, self.pasta_testes)

        # Verifica se o estado inicial é "Não Inicializado"
        self.assertEqual(self.crp.estado, "Não Inicializado")

    def test_carregar_planejamento_mrp(self):
        """
        Testa se o planejamento do MRP é carregado corretamente.
        """
        # Carrega o planejamento do MRP
        resultado = self.crp.carregar_planejamento_mrp("planejamento_mrp.xlsx")

        # Verifica se o carregamento foi bem-sucedido
        self.assertTrue(resultado)

        # Verifica se o estado mudou para "Inicializado"
        self.assertEqual(self.crp.estado, "Inicializado")

        # Verifica se o dicionário de planejamento foi preenchido corretamente
        planejamento = self.crp.planejamento_mrp
        self.assertIn("ETF", planejamento)
        self.assertIn("ETI", planejamento)
        self.assertIn("JOKER", planejamento)
        self.assertIn("DAQ", planejamento)

        # Verifica os valores do estoque atual
        self.assertEqual(planejamento["ETF"]["Estoque Atual"], 5)
        self.assertEqual(planejamento["ETI"]["Estoque Atual"], 10)

        # Verifica os valores das datas de entrega
        self.assertEqual(planejamento["ETF"]["2023-10-01"], 100)
        self.assertEqual(planejamento["ETF"]["2023-10-15"], 50)
        self.assertEqual(planejamento["ETF"]["2023-10-30"], 75)

        # Verifica se as datas foram extraídas corretamente
        self.assertIn("2023-10-01", self.crp.datas_entrega)
        self.assertIn("2023-10-15", self.crp.datas_entrega)
        self.assertIn("2023-10-30", self.crp.datas_entrega)

        # Verifica se os produtos foram extraídos corretamente
        self.assertIn("ETF", self.crp.produtos)
        self.assertIn("ETI", self.crp.produtos)
        self.assertIn("JOKER", self.crp.produtos)
        self.assertIn("DAQ", self.crp.produtos)

    def criar_arquivo_demanda_recursos_teste(self):
        """
        Cria um arquivo de demanda por recursos de teste.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "OP1", "OP2"])
        ws.append(["ETI", 20, 40])
        ws.append(["ETF", 30, 30])
        wb.save(os.path.join(self.pasta_testes, "demanda_recursos.xlsx"))

    def test_carregar_demanda_recursos(self):
        """
        Testa se a demanda por recursos é carregada corretamente.
        """
        self.criar_arquivo_demanda_recursos_teste()
        resultado = self.crp.carregar_demanda_recursos("demanda_recursos.xlsx")
        self.assertTrue(resultado)
        self.assertIn("ETI", self.crp.demanda_recursos)
        self.assertIn("ETF", self.crp.demanda_recursos)
        self.assertEqual(self.crp.demanda_recursos["ETI"]["OP1"], 20)
        self.assertEqual(self.crp.demanda_recursos["ETI"]["OP2"], 40)
        self.assertEqual(self.crp.demanda_recursos["ETF"]["OP1"], 30)
        self.assertEqual(self.crp.demanda_recursos["ETF"]["OP2"], 30)

    def test_calcular_demanda_por_operacao(self):
        """
        Testa o cálculo da demanda por operação.
        """
        # Configurar dados de teste
        self.crp.planejamento_mrp = {
            "ETI": {"Estoque Atual": 5, "2025-04-18": 100},
            "ETF": {"Estoque Atual": 0, "2025-04-18": 155},
        }
        self.crp.demanda_recursos = {
            "ETI": {"OP1": 20, "OP2": 40},
            "ETF": {"OP1": 30, "OP2": 30},
        }

        # Executar o método
        demanda_por_operacao = self.crp.calcular_demanda_por_operacao()

        # Verificar os resultados
        self.assertEqual(demanda_por_operacao["OP1"]["ETI"], 2000)
        self.assertEqual(demanda_por_operacao["OP1"]["ETF"], 4650)
        self.assertEqual(demanda_por_operacao["OP2"]["ETI"], 4000)
        self.assertEqual(demanda_por_operacao["OP2"]["ETF"], 4650)

        # Verificar se não há produtos extras
        self.assertEqual(len(demanda_por_operacao["OP1"]), 2)
        self.assertEqual(len(demanda_por_operacao["OP2"]), 2)

        # Verificar se não há operações extras
        self.assertEqual(len(demanda_por_operacao), 2)


    def criar_arquivo_capacidade_recursos_teste(self):
        """
        Cria um arquivo de capacidade de recursos de teste.
        """
        wb = Workbook()
        ws = wb.active

        # Adicionar cabeçalhos (primeira coluna é o recurso, demais são operações)
        ws.append(["Recurso", "OP1", "OP2", "OP3"])

        # Adicionar dados
        ws.append(["RE1", 480, 0, 480])
        ws.append(["RE2", 0, 480, 480])

        # Salvar o arquivo
        wb.save(os.path.join(self.pasta_testes, "capacidade_recursos.xlsx"))


    def test_carregar_capacidade_recursos(self):
        """
        Testa se a capacidade de recursos é carregada corretamente.
        """
        # Criar o arquivo de teste
        self.criar_arquivo_capacidade_recursos_teste()

        # Carregar a capacidade de recursos
        resultado = self.crp.carregar_capacidade_recursos("capacidade_recursos.xlsx")

        # Verificar se o carregamento foi bem-sucedido
        self.assertTrue(resultado)

        # Verificar se os recursos foram carregados corretamente
        self.assertIn("RE1", self.crp.capacidade_recursos)
        self.assertIn("RE2", self.crp.capacidade_recursos)

        # Verificar se as operações foram carregadas corretamente
        self.assertIn("OP1", self.crp.capacidade_recursos["RE1"])
        self.assertIn("OP2", self.crp.capacidade_recursos["RE1"])
        self.assertIn("OP3", self.crp.capacidade_recursos["RE1"])

        # Verificar os valores de capacidade
        self.assertEqual(self.crp.capacidade_recursos["RE1"]["OP1"], 480)
        self.assertEqual(self.crp.capacidade_recursos["RE1"]["OP2"], 0)
        self.assertEqual(self.crp.capacidade_recursos["RE1"]["OP3"], 480)
        self.assertEqual(self.crp.capacidade_recursos["RE2"]["OP1"], 0)
        self.assertEqual(self.crp.capacidade_recursos["RE2"]["OP2"], 480)
        self.assertEqual(self.crp.capacidade_recursos["RE2"]["OP3"], 480)

    def criar_arquivo_excecoes_capacidade_teste(self):
        """
        Cria um arquivo de exceções de capacidade de teste.
        """
        wb = Workbook()
        ws = wb.active

        # Adicionar dados para RE1
        ws.append(["RE1", "OP1", "OP3"])
        ws.append(["2025-03-27", 120, 120])
        ws.append(["2025-03-28", 60, 60])

        # Adicionar dados para RE2
        ws.append(["RE2", "OP2", "OP3"])
        ws.append(["2025-03-27", 30, 30])
        ws.append(["2025-03-28", 180, 180])

        # Salvar o arquivo
        wb.save(os.path.join(self.pasta_testes, "excecoes_capacidade.xlsx"))

    def test_carregar_excecoes_capacidade(self):
        """
        Testa se as exceções de capacidade são carregadas corretamente.
        """
        # Criar o arquivo de teste
        self.criar_arquivo_excecoes_capacidade_teste()

        # Carregar as exceções de capacidade
        resultado = self.crp.carregar_excecoes_capacidade("excecoes_capacidade.xlsx")

        # Verificar se o carregamento foi bem-sucedido
        self.assertTrue(resultado)

        # Verificar se os recursos foram carregados corretamente
        self.assertIn("RE1", self.crp.excecoes_capacidade)
        self.assertIn("RE2", self.crp.excecoes_capacidade)

        # Verificar se as operações foram carregadas corretamente
        self.assertIn("OP1", self.crp.excecoes_capacidade["RE1"])
        self.assertIn("OP3", self.crp.excecoes_capacidade["RE1"])
        self.assertIn("OP2", self.crp.excecoes_capacidade["RE2"])
        self.assertIn("OP3", self.crp.excecoes_capacidade["RE2"])

        # Verificar se as datas foram carregadas corretamente
        self.assertIn("2025-03-27", self.crp.excecoes_capacidade["RE1"]["OP1"])
        self.assertIn("2025-03-28", self.crp.excecoes_capacidade["RE1"]["OP1"])

        # Verificar os valores de redução de capacidade
        self.assertEqual(self.crp.excecoes_capacidade["RE1"]["OP1"]["2025-03-27"], 120)
        self.assertEqual(self.crp.excecoes_capacidade["RE1"]["OP3"]["2025-03-28"], 60)
        self.assertEqual(self.crp.excecoes_capacidade["RE2"]["OP2"]["2025-03-27"], 30)
        self.assertEqual(self.crp.excecoes_capacidade["RE2"]["OP3"]["2025-03-28"], 180)

    def test_criar_planilha_crp(self):
        """
        Testa se a planilha CRP é criada corretamente.
        """
        from datetime import datetime, timedelta

        # Criar arquivos necessários para o teste
        self.criar_arquivo_planejamento_teste()
        self.criar_arquivo_demanda_recursos_teste()
        self.criar_arquivo_capacidade_recursos_teste()
        self.criar_arquivo_excecoes_capacidade_teste()

        # Carregar os dados necessários
        self.crp.carregar_planejamento_mrp("planejamento_mrp.xlsx")
        self.crp.carregar_demanda_recursos("demanda_recursos.xlsx")
        self.crp.carregar_capacidade_recursos("capacidade_recursos.xlsx")
        self.crp.carregar_excecoes_capacidade("excecoes_capacidade.xlsx")

        # Definir data de planejamento e número de dias
        data_planejamento = "2025-04-01"
        numero_dias = 3

        # Criar a planilha CRP
        arquivo_crp = "crp_teste.xlsx"
        caminho_crp = self.crp.criar_planilha_crp(arquivo_crp, data_planejamento, numero_dias)

        # Verificar se o arquivo foi criado
        self.assertTrue(os.path.exists(caminho_crp))

        # Carregar o arquivo para verificar seu conteúdo
        import openpyxl
        wb = openpyxl.load_workbook(caminho_crp)

        # Verificar se a aba de demanda total existe
        self.assertIn("Demanda Total", wb.sheetnames)

        # Verificar se as abas de alocação diária existem
        for dia in range(numero_dias):
            data_aba = (datetime.strptime(data_planejamento, '%Y-%m-%d') + timedelta(days=dia)).strftime('%Y-%m-%d')
            nome_aba = f"Alocação {data_aba}"
            self.assertIn(nome_aba, wb.sheetnames)

        # Limpar o arquivo criado
        wb.close()
        os.remove(caminho_crp)


if __name__ == '__main__':
    unittest.main()

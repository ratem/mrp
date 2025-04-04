import unittest
import os
from mrp import MRP  # Importa a classe MRP
from openpyxl import Workbook


class TestMRP(unittest.TestCase):
    """
    Classe de teste para o MRP.
    """

    def setUp(self):
        """
        Método que é executado antes de cada teste.
        Configura o ambiente de teste, criando uma instância da classe MRP.
        """
        self.pasta_testes = "test_data"  # Diretório com os arquivos de teste
        self.mrp = MRP(self.pasta_testes)
        # Cria o diretório de testes se não existir
        if not os.path.exists(self.pasta_testes):
            os.makedirs(self.pasta_testes)
        # Cria arquivos de BOM de exemplo para teste
        self.criar_arquivos_bom_teste()
        # Cria arquivo de Estoque de exemplo para teste
        self.criar_arquivo_estoque_teste()

    def tearDown(self):
        """
        Método que é executado após cada teste.
        Limpa o ambiente de teste, removendo os arquivos e diretórios criados.
        """
        # Remove o diretório de testes e seu conteúdo
        for arquivo in os.listdir(self.pasta_testes):
            caminho_arquivo = os.path.join(self.pasta_testes, arquivo)
            os.remove(caminho_arquivo)
        os.rmdir(self.pasta_testes)

    def criar_arquivos_bom_teste(self):
        """
        Cria arquivos de BOM de exemplo (XLSX) para serem usados nos testes.
        """
        # Criar arquivo ETI_BOM.xlsx
        workbook_eti = Workbook()
        sheet_eti = workbook_eti.active
        sheet_eti.append(["Material", "Quantidade"])  # Cabeçalho
        sheet_eti.append(["JOKER", 1])
        sheet_eti.append(["DAQ", 1])
        sheet_eti.append(["SX1276 (LoRa)", 1])
        sheet_eti.append(["ADS1115", 2])
        sheet_eti.append(["Conector DB25 fêmea", 1])
        sheet_eti.append(["Soquetes 2x6p", 1])
        sheet_eti.append(["Invólucro", 1])
        sheet_eti.append(["Antena LoRa", 1])
        sheet_eti.append(["D-SUB25 macho", 1])
        workbook_eti.save(os.path.join(self.pasta_testes, "ETI_BOM.xlsx"))
        # Criar arquivo ETF_BOM.xlsx
        workbook_etf = Workbook()
        sheet_etf = workbook_etf.active
        sheet_etf.append(["Material", "Quantidade"])  # Cabeçalho
        sheet_etf.append(["JOKER", 1])
        sheet_etf.append(["DAQ", 1])
        sheet_etf.append(["Módulo 3G/4G", 1])
        sheet_etf.append(["Chip Internet Móvel", 1])
        sheet_etf.append(["SX1276 (LoRa)", 1])
        sheet_etf.append(["ADS1115", 2])
        sheet_etf.append(["Conector DB25 fêmea", 1])
        sheet_etf.append(["Soquetes 2x6p", 1])
        sheet_etf.append(["Invólucro", 1])
        sheet_etf.append(["Antena LoRa", 1])
        sheet_etf.append(["D-SUB25 macho", 1])
        workbook_etf.save(os.path.join(self.pasta_testes, "ETF_BOM.xlsx"))

    def test_carregar_bom(self):
        """
        Testa se as BOMs são carregadas corretamente.
        """
        self.mrp.inicializar_dados()
        # Verifica se as BOMs foram carregadas corretamente
        self.assertIn("ETI", self.mrp.boms)
        self.assertIn("ETF", self.mrp.boms)
        self.assertEqual(self.mrp.boms["ETI"]["JOKER"], 1)
        self.assertEqual(self.mrp.boms["ETF"]["JOKER"], 1)

    def criar_arquivo_estoque_teste(self):
        """
        Cria arquivo de Estoque de exemplo (XLSX) para ser usado nos testes.
        """
        # Criar arquivo Estoque.xlsx
        workbook_estoque = Workbook()
        sheet_estoque = workbook_estoque.active
        # Dados de Estoque (obtidos da planilha anexa)
        dados_estoque = [
            ["Material", "Em Estoque", "Minimo", "Custo Medio Unitario", "Imposto Medio Unitario", "Frete Medio Lote",
             "Leadtime Medio Lote"],
            ["ETF", 0, 5, 100, 0, 0, 5],
            ["JOKER", 10, 10, 10, 1, 5, 10],
            ["DAQ", 10, 10, 11, 2, 6, 11],
            ["Módulo 3G/4G", 10, 10, 12, 3, 7, 12],
            ["Chip Internet Móvel", 5, 5, 13, 4, 8, 13],
            ["SX1276 (LoRa)", 20, 20, 14, 5, 9, 14],
            ["ADS1115", 10, 5, 15, 6, 10, 15],
            ["Conector DB25 fêmea", 30, 10, 16, 7, 11, 16],
            ["Soquetes 2x6p", 20, 20, 17, 8, 12, 17],
            ["Invólucro", 5, 5, 18, 9, 13, 18],
            ["Antena LoRa", 15, 10, 19, 10, 14, 19],
            ["D-SUB25 macho", 12, 50, 20, 11, 15, 20],
            ["ETI", 5, 5, 80, 0, 0, 5]
        ]
        # Escrever os dados na planilha
        for linha in dados_estoque:
            sheet_estoque.append(linha)
        # Salvar o arquivo
        workbook_estoque.save(os.path.join(self.pasta_testes, "Estoque.xlsx"))

    def test_calcular_quantidades_producao_aquisicao(self):
        """
        Testa o cálculo das quantidades de produção e aquisição.
        """
        # Define a demanda
        demanda = {"ETI": 100, "ETF": 100}
        # Inicializa os dados do MRP
        self.mrp.inicializar_dados()
        # Calcula as quantidades de produção e aquisição
        quantidades, ordens_planejamento = self.mrp.calcular_quantidades_producao_aquisicao(demanda)
        # Verificações
        # Produto ETI
        self.assertEqual(quantidades["ETI"]["quantidade_a_produzir"], 100)
        self.assertEqual(ordens_planejamento["ETI"]["Produção"], 100)
        # Produto ETF
        self.assertEqual(quantidades["ETF"]["quantidade_a_produzir"], 105)
        self.assertEqual(ordens_planejamento["ETF"]["Produção"], 105)
        # Componentes
        # Estoque mínimo de JOKER é 10, estoque inicial é 10. Demanda é 205. Precisa adquirir 205.
        self.assertEqual(quantidades["JOKER"]["quantidade_a_adquirir"], 205)
        self.assertEqual(ordens_planejamento["JOKER"]["Aquisição"], 205)
        # Estoque mínimo de DAQ é 10, estoque inicial é 10. Demanda é 205. Precisa adquirir 205.
        self.assertEqual(quantidades["DAQ"]["quantidade_a_adquirir"], 205)
        self.assertEqual(ordens_planejamento["DAQ"]["Aquisição"], 205)
        # Estoque mínimo de SX1276 (LoRa) é 20, estoque inicial é 20. Demanda é 205. Precisa adquirir 205.
        self.assertEqual(quantidades["SX1276 (LoRa)"]["quantidade_a_adquirir"], 205)
        self.assertEqual(ordens_planejamento["SX1276 (LoRa)"]["Aquisição"], 205)
        # Estoque mínimo de ADS1115 é 5, estoque inicial é 10. Demanda é 410. Precisa adquirir 405.
        self.assertEqual(quantidades["ADS1115"]["quantidade_a_adquirir"], 405)
        self.assertEqual(ordens_planejamento["ADS1115"]["Aquisição"], 405)
        # Estoque mínimo de Conector DB25 fêmea é 10, estoque inicial é 30. Demanda é 205. Precisa adquirir 185.
        self.assertEqual(quantidades["Conector DB25 fêmea"]["quantidade_a_adquirir"], 185)
        self.assertEqual(ordens_planejamento["Conector DB25 fêmea"]["Aquisição"], 185)
        # Estoque mínimo de Soquetes 2x6p é 20, estoque inicial é 20. Demanda é 205. Precisa adquirir 205.
        self.assertEqual(quantidades["Soquetes 2x6p"]["quantidade_a_adquirir"], 205)
        self.assertEqual(ordens_planejamento["Soquetes 2x6p"]["Aquisição"], 205)
        # Estoque mínimo de Invólucro é 5, estoque inicial é 5. Demanda é 205. Precisa adquirir 205.
        self.assertEqual(quantidades["Invólucro"]["quantidade_a_adquirir"], 205)
        self.assertEqual(ordens_planejamento["Invólucro"]["Aquisição"], 205)
        # Estoque mínimo de Antena LoRa é 10, estoque inicial é 15. Demanda é 205. Precisa adquirir 200.
        self.assertEqual(quantidades["Antena LoRa"]["quantidade_a_adquirir"], 200)
        self.assertEqual(ordens_planejamento["Antena LoRa"]["Aquisição"], 200)
        # Estoque mínimo de D-SUB25 macho é 50, estoque inicial é 12. Demanda é 205. Precisa adquirir 248.
        self.assertEqual(quantidades["D-SUB25 macho"]["quantidade_a_adquirir"], 243)
        self.assertEqual(ordens_planejamento["D-SUB25 macho"]["Aquisição"], 243)
        # Estoque mínimo de Módulo 3G/4G é 10, estoque inicial é 10. Demanda é 105. Precisa adquirir 105.
        self.assertEqual(quantidades["Módulo 3G/4G"]["quantidade_a_adquirir"], 105)
        self.assertEqual(ordens_planejamento["Módulo 3G/4G"]["Aquisição"], 105)
        # Estoque mínimo de Chip Internet Móvel é 5, estoque inicial é 5. Demanda é 105. Precisa adquirir 105.
        self.assertEqual(quantidades["Chip Internet Móvel"]["quantidade_a_adquirir"], 105)
        self.assertEqual(ordens_planejamento["Chip Internet Móvel"]["Aquisição"], 105)

    def test_calcular_fc_lt_esperados(self):
        """
        Testa o cálculo do fluxo de caixa e lead times esperados.
        """
        # Define a demanda
        demanda = {"ETI": 100, "ETF": 100}
        # Inicializa os dados do MRP
        self.mrp.inicializar_dados()
        # Calcula as quantidades de produção e aquisição
        self.mrp.calcular_quantidades_producao_aquisicao(
            demanda)  # Garante que self.ordens_planejamento esteja preenchido
        fc_lt_esperados = self.mrp.calcular_fc_lt_esperados()
        # Verificações
        # Produto ETI
        self.assertIn("ETI", fc_lt_esperados)
        self.assertIn("Leadtime", fc_lt_esperados["ETI"])
        self.assertIn("Custo", fc_lt_esperados["ETI"])
        self.assertEqual(fc_lt_esperados["ETI"]["Leadtime"], 25)
        self.assertAlmostEqual(fc_lt_esperados["ETI"]["Custo"], 8000)  # Verifique o custo
        # Produto ETF
        self.assertIn("ETF", fc_lt_esperados)
        self.assertIn("Leadtime", fc_lt_esperados["ETF"])
        self.assertIn("Custo", fc_lt_esperados["ETF"])
        self.assertEqual(fc_lt_esperados["ETF"]["Leadtime"], 25)
        self.assertAlmostEqual(fc_lt_esperados["ETF"]["Custo"], 10500)
        # Componente JOKER
        self.assertIn("JOKER", fc_lt_esperados)
        self.assertIn("Leadtime", fc_lt_esperados["JOKER"])
        self.assertIn("Custo", fc_lt_esperados["JOKER"])
        self.assertEqual(fc_lt_esperados["JOKER"]["Leadtime"], 10)
        self.assertAlmostEqual(fc_lt_esperados["JOKER"]["Custo"], 2260)
        # Componente DAQ
        self.assertIn("DAQ", fc_lt_esperados)
        self.assertIn("Leadtime", fc_lt_esperados["DAQ"])
        self.assertIn("Custo", fc_lt_esperados["DAQ"])
        self.assertEqual(fc_lt_esperados["DAQ"]["Leadtime"], 11)
        self.assertAlmostEqual(fc_lt_esperados["DAQ"]["Custo"], 2671)
        # Componente Módulo 3G/4G
        self.assertIn("Módulo 3G/4G", fc_lt_esperados)
        self.assertEqual(fc_lt_esperados["Módulo 3G/4G"]["Leadtime"], 12)
        self.assertAlmostEqual(fc_lt_esperados["Módulo 3G/4G"]["Custo"], 1582)
        # Componente Chip Internet Móvel
        self.assertIn("Chip Internet Móvel", fc_lt_esperados)
        self.assertEqual(fc_lt_esperados["Chip Internet Móvel"]["Leadtime"], 13)
        self.assertAlmostEqual(fc_lt_esperados["Chip Internet Móvel"]["Custo"], 1793)
        # Componente SX1276 (LoRa)
        self.assertIn("SX1276 (LoRa)", fc_lt_esperados)
        self.assertEqual(fc_lt_esperados["SX1276 (LoRa)"]["Leadtime"], 14)
        self.assertAlmostEqual(fc_lt_esperados["SX1276 (LoRa)"]["Custo"], 3904)
        # Componente ADS1115
        self.assertIn("ADS1115", fc_lt_esperados)
        self.assertEqual(fc_lt_esperados["ADS1115"]["Leadtime"], 15)
        self.assertAlmostEqual(fc_lt_esperados["ADS1115"]["Custo"], 8515)
        # Componente Conector DB25 fêmea
        self.assertIn("Conector DB25 fêmea", fc_lt_esperados)
        self.assertEqual(fc_lt_esperados["Conector DB25 fêmea"]["Leadtime"], 16)
        self.assertAlmostEqual(fc_lt_esperados["Conector DB25 fêmea"]["Custo"], 4266)
        # Componente Soquetes 2x6p
        self.assertIn("Soquetes 2x6p", fc_lt_esperados)
        self.assertEqual(fc_lt_esperados["Soquetes 2x6p"]["Leadtime"], 17)
        self.assertAlmostEqual(fc_lt_esperados["Soquetes 2x6p"]["Custo"], 5137)
        # Componente Invólucro
        self.assertIn("Invólucro", fc_lt_esperados)
        self.assertEqual(fc_lt_esperados["Invólucro"]["Leadtime"], 18)
        self.assertAlmostEqual(fc_lt_esperados["Invólucro"]["Custo"], 5548)
        # Componente Antena LoRa
        self.assertIn("Antena LoRa", fc_lt_esperados)
        self.assertEqual(fc_lt_esperados["Antena LoRa"]["Leadtime"], 19)
        self.assertAlmostEqual(fc_lt_esperados["Antena LoRa"]["Custo"], 5814)
        # Componente D-SUB25 macho
        self.assertIn("D-SUB25 macho", fc_lt_esperados)
        self.assertEqual(fc_lt_esperados["D-SUB25 macho"]["Leadtime"], 20)
        self.assertAlmostEqual(fc_lt_esperados["D-SUB25 macho"]["Custo"], 7548)

    def test_montar_quadro_planejamento(self):
        """
        Testa o método montar_quadro_planejamento.
        """
        from datetime import datetime, timedelta

        # Define a demanda
        demanda = {"ETI": 100, "ETF": 100}
        # Inicializa os dados do MRP
        self.mrp.inicializar_dados()
        # Calcula as quantidades de produção e aquisição
        self.mrp.calcular_quantidades_producao_aquisicao(demanda)
        # Calcula leadtimes
        self.mrp.calcular_fc_lt_esperados()
        # Monta o quadro de planejamento
        quadro_planejamento = self.mrp.montar_quadro_planejamento()
        # Verificações
        # Verifica se o quadro de planejamento foi criado corretamente
        self.assertIn("ETI", quadro_planejamento)
        self.assertIn("ETF", quadro_planejamento)
        self.assertIn("JOKER", quadro_planejamento)
        # Verifica se os valores estão corretos (ajuste conforme necessário)
        self.assertEqual(quadro_planejamento["ETI"]["Estoque Atual"], 5)
        self.assertEqual(quadro_planejamento["ETF"]["Estoque Atual"], 0)
        self.assertEqual(quadro_planejamento["JOKER"]["Estoque Atual"], 10)
        # Verifica se as datas de entrega foram calculadas corretamente
        # Lead time de ETI: 5 (produto) + max(10, 11, 14, 16, 17, 18, 19, 20)
        data_entrega_eti = datetime.now() + timedelta(days=5 + 20)
        data_entrega_eti_str = data_entrega_eti.strftime('%Y-%m-%d')
        self.assertIn(data_entrega_eti_str, quadro_planejamento["ETI"])
        self.assertEqual(quadro_planejamento["ETI"][data_entrega_eti_str], 100)  # Quantidade de ETI
        # Lead time de ETF: 5 (produto) + max(10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20) = 20
        data_entrega_etf = datetime.now() + timedelta(days=5 + 20)
        data_entrega_etf_str = data_entrega_etf.strftime('%Y-%m-%d')
        self.assertIn(data_entrega_etf_str, quadro_planejamento["ETF"])
        self.assertEqual(quadro_planejamento["ETF"][data_entrega_etf_str], 105)  # Quantidade de ETF

    def test_iniciar_execucao(self):
        """
        Testa se a execução é iniciada corretamente.
        """
        # Define a demanda
        demanda = {"ETI": 100, "ETF": 100}
        # Inicializa os dados do MRP
        self.mrp.inicializar_dados()
        # Realiza o planejamento
        self.mrp.planejar_producao(demanda)
        self.assertEqual(self.mrp.estado, "Planejado")
        resultado = self.mrp.iniciar_execucao()
        # Verifica se a execução foi iniciada com sucesso
        self.assertTrue(resultado)
        # Verifica se o estado mudou para Em Execução
        self.assertEqual(self.mrp.estado, "Em Execução")
        # Verifica se a execução foi iniciada com sucesso
        self.assertTrue(resultado)
        # Verifica se o estado mudou para Em Execução
        self.assertEqual(self.mrp.estado, "Em Execução")
        # Verifica se o dicionário ordens_controle foi criado
        self.assertIsNotNone(self.mrp.ordens_controle)
        # Verifica se todas as ordens de planejamento foram copiadas para ordens_controle
        for material in self.mrp.ordens_planejamento:
            self.assertIn(material, self.mrp.ordens_controle)
            # Verifica se o status inicial é 'Planejada'
            self.assertEqual(self.mrp.ordens_controle[material]['Status'], 'Planejada')
            # Verifica se o estoque atual foi copiado corretamente
            estoque_esperado = self.mrp.estoque.get(material, {'em_estoque': 0})['em_estoque']
            self.assertEqual(self.mrp.ordens_controle[material]['Estoque Atual'], estoque_esperado)
            # Verifica se as ordens de produção foram copiadas corretamente
            if 'Produção' in self.mrp.ordens_planejamento[material]:
                self.assertIn('Produção', self.mrp.ordens_controle[material])
                self.assertEqual(
                    self.mrp.ordens_controle[material]['Produção'],
                    self.mrp.ordens_planejamento[material]['Produção']
                )
            # Verifica se as ordens de aquisição foram copiadas corretamente
            if 'Aquisição' in self.mrp.ordens_planejamento[material]:
                self.assertEqual(
                    self.mrp.ordens_controle[material].get('Aquisição'),
                    self.mrp.ordens_planejamento[material]['Aquisição']
                )

    def test_editar_cancelar_ordem(self):
        """
        Testa se é possível editar e cancelar ordens.
        A sequência de eventos deve ser:
        "Ordem de Aquisição para 'JOKER' editada: 205 -> 255"
        Uma ordem de aquisição para o componente 'JOKER' foi editada, aumentando a quantidade de 205 para 255 unidades.
        "Status da ordem para 'JOKER' atualizado para 'Executada'."
        O status da ordem foi alterado para "Executada", indicando que a ordem está em processamento.
        "Status da ordem para 'JOKER' atualizado para 'Planejada' devido à edição."
        Quando a ordem foi editada novamente, o sistema automaticamente reverteu o status para "Planejada", seguindo a regra de que ordens editadas devem voltar ao estado inicial.
        "Ordem de Aquisição para 'JOKER' editada: 255 -> 235"
        A quantidade da ordem foi reduzida de 255 para 235 unidades.
        "Ordem para 'JOKER' cancelada. Quantidades zeradas e status definido como 'Planejada'."
        A ordem foi completamente cancelada, zerando as quantidades e definindo o status como "Planejada".
        "Erro: Não é possível atualizar o status de uma ordem cancelada (quantidades zeradas)."
        Uma tentativa de atualizar o status de uma ordem cancelada foi rejeitada pelo sistema, conforme esperado.
        "Execução iniciada com sucesso. Estado atual: Em Execução"
        O MRP iniciou uma nova fase de execução.
        """
        # Define a demanda
        demanda = {"ETI": 100, "ETF": 100}
        # Inicializa os dados do MRP
        self.mrp.inicializar_dados()
        # Realiza o planejamento
        self.mrp.planejar_producao(demanda)
        # Inicia a execução
        self.mrp.iniciar_execucao()
        # Verifica se o estado mudou para Em Execução
        self.assertEqual(self.mrp.estado, "Em Execução")
        # Seleciona um material para testar (JOKER)
        material = "JOKER"
        # Guarda a quantidade inicial de aquisição
        quantidade_inicial = self.mrp.ordens_controle[material]['Aquisição']
        # Edita a ordem
        nova_quantidade = quantidade_inicial + 50
        resultado = self.mrp.editar_ordem(material, 'Aquisição', nova_quantidade)
        self.assertTrue(resultado)
        # Verifica se a quantidade foi atualizada
        self.assertEqual(self.mrp.ordens_controle[material]['Aquisição'], nova_quantidade)
        # Verifica se o status permanece como 'Planejada'
        self.assertEqual(self.mrp.ordens_controle[material]['Status'], 'Planejada')
        # Atualiza o status para 'Executada'
        self.mrp.atualizar_status_ordem(material, 'Executada')
        # Edita a ordem novamente
        nova_quantidade_2 = nova_quantidade - 20
        resultado = self.mrp.editar_ordem(material, 'Aquisição', nova_quantidade_2)
        self.assertTrue(resultado)
        # Verifica se a quantidade foi atualizada
        self.assertEqual(self.mrp.ordens_controle[material]['Aquisição'], nova_quantidade_2)
        # Verifica se o status voltou para 'Planejada'
        self.assertEqual(self.mrp.ordens_controle[material]['Status'], 'Planejada')
        # Testa o cancelamento de ordem
        resultado = self.mrp.cancelar_ordem(material)
        self.assertTrue(resultado)
        # Verifica se as quantidades foram zeradas
        self.assertEqual(self.mrp.ordens_controle[material]['Aquisição'], 0)
        # Verifica se o status é 'Planejada'
        self.assertEqual(self.mrp.ordens_controle[material]['Status'], 'Planejada')
        # Tenta atualizar o status de uma ordem cancelada
        resultado = self.mrp.atualizar_status_ordem(material, 'Executada')
        self.assertFalse(resultado)

    def test_atualizar_status_ordem(self):
        """
        Testa se o status de uma ordem é atualizado corretamente.
        A sequência de eventos no teste deve ser:
        Inicialização da execução: O MRP mudou seu estado para "Em Execução" com sucesso.
        Atualização para 'Executada': O status da ordem para o material 'JOKER' foi atualizado para 'Executada' com sucesso.
        Atualização para 'Pronta': O status da ordem para 'JOKER' foi atualizado para 'Pronta' com sucesso, o que desencadeou a atualização do estoque.
        Atualização do estoque: Como 'JOKER' não tinha ordem de produção (apenas aquisição), o sistema corretamente adicionou 0 unidades de produção e 205 unidades de aquisição ao estoque.
        Teste de casos inválidos: O sistema rejeitou corretamente:
        Um status inválido ("Status Inválido")
        Um material inexistente ("Material Inexistente")
        Uma atualização quando o MRP não estava em execução (após mudar o estado)
        Nova inicialização: No final, o teste parece ter iniciado a execução novamente.
        """
        # Define a demanda
        demanda = {"ETI": 100, "ETF": 100}
        # Inicializa os dados do MRP
        self.mrp.inicializar_dados()
        # Define o estado como Inicializado para poder planejar
        self.mrp.estado = "Inicializado"
        # Realiza o planejamento
        self.mrp.planejar_producao(demanda)
        # Inicia a execução
        self.mrp.iniciar_execucao()
        # Verifica se o estado mudou para Em Execução
        self.assertEqual(self.mrp.estado, "Em Execução")
        # Seleciona um material para testar (JOKER)
        material = "JOKER"
        # Guarda o estoque inicial
        estoque_inicial = self.mrp.estoque[material]['em_estoque']
        # Verifica se o status inicial é 'Planejada'
        self.assertEqual(self.mrp.ordens_controle[material]['Status'], 'Planejada')
        # Atualiza o status para 'Executada'
        resultado = self.mrp.atualizar_status_ordem(material, 'Executada')
        self.assertTrue(resultado)
        self.assertEqual(self.mrp.ordens_controle[material]['Status'], 'Executada')
        # Verifica que o estoque não mudou
        self.assertEqual(self.mrp.estoque[material]['em_estoque'], estoque_inicial)
        # Atualiza o status para 'Pronta'
        resultado = self.mrp.atualizar_status_ordem(material, 'Pronta')
        self.assertTrue(resultado)
        self.assertEqual(self.mrp.ordens_controle[material]['Status'], 'Pronta')
        # Verifica que o estoque foi atualizado
        quantidade_aquisicao = self.mrp.ordens_controle[material]['Aquisição']
        self.assertEqual(self.mrp.estoque[material]['em_estoque'], estoque_inicial + quantidade_aquisicao)
        # Testa com status inválido
        resultado = self.mrp.atualizar_status_ordem(material, 'Status Inválido')
        self.assertFalse(resultado)
        # Testa com material inexistente
        resultado = self.mrp.atualizar_status_ordem('Material Inexistente', 'Pronta')
        self.assertFalse(resultado)
        # Testa quando o MRP não está em execução
        self.mrp.estado = "Planejado"
        resultado = self.mrp.atualizar_status_ordem(material, 'Pronta')
        self.assertFalse(resultado)

    def criar_arquivo_cotacoes_teste(self):
        """
        Cria um arquivo de cotações de teste para ser usado nos testes.
        """
        from openpyxl import Workbook

        # Criar arquivo Cotacoes.xlsx
        workbook_cotacoes = Workbook()
        sheet_cotacoes = workbook_cotacoes.active
        sheet_cotacoes.append(["Material", "Custo Unitario", "Imposto Unitario", "Frete Lote", "Lead Time"])
        sheet_cotacoes.append(["ETF", 70.0, 15.0, 0.0, 5])
        sheet_cotacoes.append(["JOKER", 42.41, 47.68, 6.58, 15])
        sheet_cotacoes.append(["DAQ", 17.19, 29.22, 11.46, 15])

        # Salvar o arquivo
        workbook_cotacoes.save(os.path.join(self.pasta_testes, "Cotacoes.xlsx"))

    def test_atualizar_custos_leadtimes(self):
        """
        Testa a atualização de custos e leadtimes a partir da planilha de cotações.
        """
        # Define a demanda
        demanda = {"ETI": 100, "ETF": 100}

        # Inicializa os dados do MRP
        self.mrp.inicializar_dados()

        # Realiza o planejamento
        self.mrp.planejar_producao(demanda)

        # Inicia a execução
        self.mrp.iniciar_execucao()

        # Verifica se o estado mudou para Em Execução
        self.assertEqual(self.mrp.estado, "Em Execução")

        # Cria uma planilha de cotações de teste
        self.criar_arquivo_cotacoes_teste()

        # Salva os valores originais para comparação
        joker_custo_original = self.mrp.estoque["JOKER"]["custo_medio_unitario"]
        joker_leadtime_original = self.mrp.estoque["JOKER"]["leadtime_medio_lote"]
        etf_custo_original = self.mrp.estoque["ETF"]["custo_medio_unitario"]

        # Atualiza os custos e leadtimes
        sucesso, alertas = self.mrp.atualizar_custos_leadtimes("Cotacoes.xlsx")

        # Verifica se a atualização foi bem-sucedida
        self.assertTrue(sucesso)

        # Verifica se foram gerados alertas
        self.assertTrue(len(alertas) > 0)

        # Verifica se os valores foram atualizados corretamente
        self.assertEqual(self.mrp.estoque["JOKER"]["custo_medio_unitario"], 42.41)
        self.assertEqual(self.mrp.estoque["JOKER"]["leadtime_medio_lote"], 15)
        self.assertEqual(self.mrp.estoque["ETF"]["custo_medio_unitario"], 70.0)

        # Verifica se os alertas contêm as informações essenciais de forma mais flexível
        # Em vez de verificar strings exatas, verificamos se os componentes essenciais estão presentes
        self.assertTrue(any("Custo de JOKER" in alerta and "10" in alerta and "42.41" in alerta for alerta in alertas),
                        "Alerta sobre alteração de custo do JOKER não encontrado")

        self.assertTrue(any(
            "Leadtime de JOKER" in alerta and str(joker_leadtime_original) in alerta and "15" in alerta for alerta in
            alertas),
                        "Alerta sobre alteração de leadtime do JOKER não encontrado")

        self.assertTrue(any("Custo de ETF" in alerta and "100" in alerta and "70" in alerta for alerta in alertas),
                        "Alerta sobre alteração de custo do ETF não encontrado")

    def criar_arquivo_planejamento_teste(self):
        """
        Cria um arquivo de planejamento de teste para ser usado nos testes.
        """
        from openpyxl import Workbook

        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active

        # Adicionar cabeçalhos
        ws.append(["Material", "Estoque Atual", "2025-03-30", "2025-04-05"])

        # Adicionar dados
        ws.append(["ETF", 5, 100, 50])
        ws.append(["JOKER", 10, 200, 100])
        ws.append(["DAQ", 15, 300, 150])

        # Salvar o arquivo
        wb.save(os.path.join(self.pasta_testes, "Planejamento_teste.xlsx"))

    def test_recuperar_planejamento(self):
        """
        Testa se o planejamento é recuperado corretamente de uma planilha Excel.
        """
        # Criar uma planilha de planejamento de teste
        self.criar_arquivo_planejamento_teste()

        # Recuperar o planejamento
        resultado = self.mrp.recuperar_planejamento("Planejamento_teste.xlsx")

        # Verificar se a recuperação foi bem-sucedida
        self.assertTrue(resultado)

        # Verificar se o dicionário de planejamento foi preenchido corretamente
        planejamento = self.mrp.planejamento
        self.assertIn("ETF", planejamento)
        self.assertIn("JOKER", planejamento)
        self.assertIn("DAQ", planejamento)

        # Verificar os valores do estoque atual
        self.assertEqual(planejamento["ETF"]["Estoque Atual"], 5)
        self.assertEqual(planejamento["JOKER"]["Estoque Atual"], 10)
        self.assertEqual(planejamento["DAQ"]["Estoque Atual"], 15)

        # Verificar os valores das datas de entrega
        self.assertEqual(planejamento["ETF"]["2025-03-30"], 100)
        self.assertEqual(planejamento["ETF"]["2025-04-05"], 50)

        self.assertEqual(planejamento["JOKER"]["2025-03-30"], 200)
        self.assertEqual(planejamento["JOKER"]["2025-04-05"], 100)

        self.assertEqual(planejamento["DAQ"]["2025-03-30"], 300)
        self.assertEqual(planejamento["DAQ"]["2025-04-05"], 150)



if __name__ == '__main__':
    unittest.main()
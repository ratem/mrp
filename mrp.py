import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta

class MRP:
    """
    Classe que implementa o Material Requirements Planning (MRP).
    """

    def __init__(self, pasta_arquivos):
        """
        Inicializa o objeto MRP com a pasta contendo os arquivos de entrada.
        """
        self.pasta_arquivos = pasta_arquivos
        self.estoque = {}
        self.boms = {}  # Bills of Materials
        self.fc_lt_esperados = {}
        self.planejamento = {}


    def carregar_estoque(self, df_estoque):
        """
        Carrega os dados de estoque a partir de um DataFrame do pandas.
        """
        for _, row in df_estoque.iterrows():
            codigo_material = row['Material']
            self.estoque[codigo_material] = {
                'em_estoque': row['Em Estoque'],
                'minimo': row['Minimo'],
                'custo_medio_unitario': row['Custo Medio Unitario'],
                'imposto_medio_unitario': row['Imposto Medio Unitario'],
                'frete_medio_lote': row['Frete Medio Lote'],
                'leadtime_medio_lote': row['Leadtime Medio Lote']
            }

    def carregar_bom(self, codigo_produto, df_bom):
        """
        Carrega a BOM (Bill of Materials) a partir de um DataFrame do pandas.
        """
        self.boms[codigo_produto] = {}
        for _, row in df_bom.iterrows():
            material = row['Material']
            quantidade = row['Quantidade']
            self.boms[codigo_produto][material] = quantidade

    def inicializar_dados(self):
        """
        Carrega os dados de estoque e BOMs a partir dos arquivos XLSX.
        """
        # Carrega o estoque
        estoque_path = os.path.join(self.pasta_arquivos, "Estoque.xlsx")
        estoque_wb = load_workbook(estoque_path)
        estoque_sheet = estoque_wb.active
        df_estoque = pd.DataFrame(estoque_sheet.values)
        df_estoque.columns = df_estoque.iloc[0]
        df_estoque = df_estoque[1:]
        df_estoque = df_estoque.dropna(how='all')  # Remove linhas completamente vazias
        self.carregar_estoque(df_estoque)

        # Carrega as BOMs
        for arquivo in os.listdir(self.pasta_arquivos):
            if "_BOM.xlsx" in arquivo:
                codigo_produto = arquivo.replace("_BOM.xlsx", "")
                bom_path = os.path.join(self.pasta_arquivos, arquivo)
                bom_wb = load_workbook(bom_path)
                bom_sheet = bom_wb.active
                df_bom = pd.DataFrame(bom_sheet.values)
                df_bom.columns = df_bom.iloc[0]
                df_bom = df_bom[1:]
                df_bom = df_bom.dropna(how='all')  # Adicionado dropna() para remover linhas completamente vazias
                self.carregar_bom(codigo_produto, df_bom)
        self.estado = "Inicializado"

    def calcular_quantidades_producao_aquisicao(self, demanda):
        """
        Calcula as quantidades de produção e aquisição para atender à demanda.
        """
        quantidades = {}
        self.ordens_planejamento = {}

        # 1. Calcular a quantidade a produzir para os produtos finais
        for produto, quantidade_demanda in demanda.items():
            # Inicializa as entradas para o produto
            if produto not in quantidades:
                quantidades[produto] = {'quantidade_a_produzir': 0, 'quantidade_a_adquirir': 0}
            if produto not in self.ordens_planejamento:
                self.ordens_planejamento[produto] = {'Produção': 0, 'Aquisição': 0}

            # Calcula a quantidade a produzir do produto final
            estoque_atual = self.estoque.get(produto, {'em_estoque': 0})['em_estoque']
            estoque_minimo = self.estoque.get(produto, {'minimo': 0})['minimo']
            quantidade_necessaria = max(0, quantidade_demanda + estoque_minimo - estoque_atual)
            quantidades[produto]['quantidade_a_produzir'] = quantidade_necessaria
            self.ordens_planejamento[produto]['Produção'] = quantidade_necessaria

        # 2. Calcular a quantidade a adquirir para os componentes
        componentes_necessarios = {}  # Dicionário para acumular as necessidades dos componentes

        for produto, quantidade_demanda in demanda.items():
            if produto in self.boms:
                bom = self.boms[produto]
                # Calcula a quantidade necessária para o produto
                estoque_atual_produto = self.estoque.get(produto, {'em_estoque': 0})['em_estoque']
                estoque_minimo_produto = self.estoque.get(produto, {'minimo': 0})['minimo']
                quantidade_necessaria = max(0, quantidade_demanda + estoque_minimo_produto - estoque_atual_produto)

                for componente, quantidade_por_produto in bom.items():
                    # Calcula a necessidade total do componente
                    necessidade_componente = quantidade_necessaria * quantidade_por_produto

                    # Acumula as necessidades dos componentes
                    if componente not in componentes_necessarios:
                        componentes_necessarios[componente] = 0
                    componentes_necessarios[componente] += necessidade_componente

        # Calcula a quantidade a adquirir para cada componente, considerando o estoque
        for componente, necessidade_total in componentes_necessarios.items():
            # Inicializa o componente no dicionário quantidades, se necessário
            if componente not in quantidades:
                quantidades[componente] = {'quantidade_a_produzir': 0, 'quantidade_a_adquirir': 0}
            if componente not in self.ordens_planejamento:
                self.ordens_planejamento[componente] = {'Produção': 0, 'Aquisição': 0}

            # Calcula a quantidade a adquirir (CORREÇÃO AQUI)
            estoque_componente = self.estoque.get(componente, {'em_estoque': 0})['em_estoque']
            estoque_minimo_componente = self.estoque.get(componente, {'minimo': 0})['minimo']
            quantidade_a_adquirir = max(0, necessidade_total + estoque_minimo_componente - estoque_componente)

            # Define a quantidade a adquirir
            quantidades[componente]['quantidade_a_adquirir'] = quantidade_a_adquirir
            self.ordens_planejamento[componente]['Aquisição'] = quantidade_a_adquirir

        return quantidades, self.ordens_planejamento


    def calcular_fc_lt_esperados(self):
        """
        Calcula o fluxo de caixa esperado e os lead times esperados para as ordens de produção e aquisição.

        Returns:
            dict: Dicionário contendo o fluxo de caixa esperado e os lead times esperados.
        """
        self.fc_lt_esperados = {}  # Inicializa o dicionário aqui

        for material, ordens in self.ordens_planejamento.items():  # Usar self.ordens_planejamento
            self.fc_lt_esperados[material] = {}

            # Lógica para Produtos (Produção)
            if 'Produção' in ordens and ordens['Produção'] > 0:
                # Aqui você precisará implementar a lógica para calcular o lead time do produto final.
                # Conforme a descrição, é o maior lead-time entre seus componentes a adquirir somado ao seu próprio lead-time de produção.
                # Por enquanto, vamos usar um valor padrão para o lead time do produto.
                # Encontrar o maior lead time dos componentes
                maior_leadtime_componente = 0
                if material in self.boms:
                    for componente in self.boms[material]:
                        leadtime_componente = self.estoque.get(componente, {'leadtime_medio_lote': 0})[
                            'leadtime_medio_lote']
                        maior_leadtime_componente = max(maior_leadtime_componente, leadtime_componente)
                # Somar o lead time do produto
                leadtime_produto = self.estoque.get(material, {'leadtime_medio_lote': 0})['leadtime_medio_lote']
                self.fc_lt_esperados[material]['Leadtime'] = leadtime_produto + maior_leadtime_componente

                # Calcular o custo total de produção
                custo_unitario = self.estoque.get(material, {'custo_medio_unitario': 0})['custo_medio_unitario']
                imposto_unitario = self.estoque.get(material, {'imposto_medio_unitario': 0})['imposto_medio_unitario']
                # Frete não se aplica à produção, então não o incluímos.
                self.fc_lt_esperados[material]['Custo'] = ordens['Produção'] * (
                            custo_unitario + imposto_unitario)

            # Lógica para Componentes (Aquisição)
            if 'Aquisição' in ordens and ordens['Aquisição'] > 0:
                # Lead time do componente é copiado do dicionário de estoque
                self.fc_lt_esperados[material]['Leadtime'] = self.estoque.get(material,
                                                                              {'leadtime_medio_lote': 0})[
                    'leadtime_medio_lote']

                # Calcular o custo total de aquisição
                custo_unitario = self.estoque.get(material, {'custo_medio_unitario': 0})['custo_medio_unitario']
                imposto_unitario = self.estoque.get(material, {'imposto_medio_unitario': 0})['imposto_medio_unitario']
                frete_lote = self.estoque.get(material, {'frete_medio_lote': 0})['frete_medio_lote']
                self.fc_lt_esperados[material]['Custo'] = ordens['Aquisição'] * (
                            custo_unitario + imposto_unitario) + frete_lote

        return self.fc_lt_esperados


    def montar_quadro_planejamento(self, data_execucao=None):
        """
        Monta o quadro de planejamento com as ordens de produção e aquisição.

        Args:
            data_execucao (datetime, opcional): Data de execução do planejamento.
                                               Se não fornecida, usa a data atual.

        Returns:
            dict: Dicionário representando o quadro de planejamento.
        """
        if data_execucao is None:
            data_execucao = datetime.now()

        for material, ordens in self.ordens_planejamento.items():
            if material not in self.planejamento:
                self.planejamento[material] = {"Estoque Atual": self.estoque.get(material, {'em_estoque': 0})['em_estoque']}
            if 'Produção' in ordens and ordens['Produção'] > 0:
                leadtime = self.fc_lt_esperados[material]['Leadtime']
                data_entrega = data_execucao + timedelta(days=leadtime)
                self.planejamento[material][data_entrega.strftime('%Y-%m-%d')] = ordens['Produção']
            if 'Aquisição' in ordens and ordens['Aquisição'] > 0:
                leadtime = self.fc_lt_esperados[material]['Leadtime']
                data_entrega = data_execucao + timedelta(days=leadtime)
                self.planejamento[material][data_entrega.strftime('%Y-%m-%d')] = ordens['Aquisição']
        self.estado = "Planejado"
        return self.planejamento


    def imprimir_quadro_planejamento(self):
        """
        Imprime o quadro de planejamento de forma tabular no console.
        """
        from tabulate import tabulate

        tabela = []
        cabecalho = ["Material", "Estoque Atual"]
        datas_entrega = set()

        # Coleta todas as datas de entrega únicas
        for material, dados in self.planejamento.items():
            for chave in dados.keys():
                if chave != "Estoque Atual":
                    datas_entrega.add(chave)

        # Adiciona as datas de entrega ao cabeçalho
        cabecalho.extend(sorted(list(datas_entrega)))

        # Popula a tabela com os dados
        for material, dados in self.planejamento.items():
            linha = [material, dados["Estoque Atual"]]
            for data in sorted(list(datas_entrega)):
                linha.append(dados.get(data, ""))  # Usa "" se a data não existir para o material
            tabela.append(linha)

        # Imprime a tabela no console
        print(tabulate(tabela, headers=cabecalho, tablefmt="grid"))

    def exportar_quadro_planejamento(self, nome_arquivo):
        """
        Exporta o quadro de planejamento para uma planilha Excel usando openpyxl.

        Args:
            nome_arquivo (str): Nome do arquivo Excel a ser criado.

        Returns:
            str: Caminho completo do arquivo Excel criado.
        """
        if not self.planejamento:
            print("Erro: O quadro de planejamento ainda não foi gerado.")
            return None

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side

        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Planejamento"

        # Coletar todas as datas de entrega únicas
        datas_entrega = set()
        for material, dados in self.planejamento.items():
            for chave in dados.keys():
                if chave != "Estoque Atual":
                    datas_entrega.add(chave)

        # Ordenar as datas
        datas_ordenadas = sorted(list(datas_entrega))

        # Criar cabeçalhos
        cabecalhos = ["Material", "Estoque Atual"] + datas_ordenadas
        for col_idx, header in enumerate(cabecalhos, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Preencher os dados
        row_idx = 2
        for material, dados in self.planejamento.items():
            ws.cell(row=row_idx, column=1, value=material)
            ws.cell(row=row_idx, column=2, value=dados["Estoque Atual"])

            # Preencher as datas para este material
            for col_idx, data in enumerate(datas_ordenadas, 3):
                if data in dados:
                    ws.cell(row=row_idx, column=col_idx, value=dados[data])

            row_idx += 1

        # Aplicar formatação
        # Definir estilo para cabeçalhos
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar estilo aos cabeçalhos
        for col_idx in range(1, len(cabecalhos) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

            # Ajustar largura da coluna
            if col_idx == 1:
                ws.column_dimensions[cell.column_letter].width = 20  # Material
            elif col_idx == 2:
                ws.column_dimensions[cell.column_letter].width = 15  # Estoque Atual
            else:
                ws.column_dimensions[cell.column_letter].width = 15  # Datas

        # Congelar a primeira linha e coluna
        ws.freeze_panes = 'B2'

        # Definir o caminho completo do arquivo
        caminho_arquivo = os.path.join(self.pasta_arquivos, nome_arquivo)

        # Salvar o arquivo
        wb.save(caminho_arquivo)

        print(f"Quadro de planejamento exportado com sucesso para: {caminho_arquivo}")
        return caminho_arquivo

    def exportar_ordens_producao(self, nome_arquivo):
        """
        Exporta as ordens de produção e aquisição para uma planilha Excel.

        Args:
            nome_arquivo (str): Nome do arquivo Excel a ser criado.

        Returns:
            str: Caminho completo do arquivo Excel criado.
        """
        if not hasattr(self, 'ordens_planejamento') or not self.ordens_planejamento:
            print("Erro: As ordens de produção ainda não foram geradas.")
            return None

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Ordens"

        # Definir cabeçalhos
        cabecalhos = [
            "Material",
            "Retirada de Estoque",
            "Produção",
            "Aquisição",
            "Custo Total Estimado",
            "Leadtime Total Estimado",
            "Custo Total Real",
            "Leadtime Total Real"
        ]

        # Adicionar cabeçalhos
        for col_idx, header in enumerate(cabecalhos, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Preencher os dados
        row_idx = 2
        for material, ordens in self.ordens_planejamento.items():
            ws.cell(row=row_idx, column=1, value=material)

            # Calcular a retirada de estoque (se houver)
            estoque_atual = self.estoque.get(material, {'em_estoque': 0})['em_estoque']

            # Produção (apenas para produtos finais)
            producao = ordens.get('Produção', 0)
            ws.cell(row=row_idx, column=3, value=producao if producao > 0 else "")

            # Aquisição (apenas para componentes)
            aquisicao = ordens.get('Aquisição', 0)
            ws.cell(row=row_idx, column=4, value=aquisicao if aquisicao > 0 else "")

            # Retirada de estoque
            retirada_estoque = min(estoque_atual, producao + aquisicao)
            ws.cell(row=row_idx, column=2, value=retirada_estoque if retirada_estoque > 0 else "")

            # Custo e Leadtime estimados
            if material in self.fc_lt_esperados:
                custo = self.fc_lt_esperados[material].get('Custo', 0)
                leadtime = self.fc_lt_esperados[material].get('Leadtime', 0)
                ws.cell(row=row_idx, column=5, value=custo)
                ws.cell(row=row_idx, column=6, value=leadtime)

            # Custo e Leadtime reais (inicialmente vazios)
            ws.cell(row=row_idx, column=7, value="")
            ws.cell(row=row_idx, column=8, value="")

            row_idx += 1

        # Aplicar formatação
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar estilo aos cabeçalhos
        for col_idx in range(1, len(cabecalhos) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

            # Ajustar largura da coluna
            ws.column_dimensions[cell.column_letter].width = 18

        # Ajustar a primeira coluna para ser mais larga (Material)
        ws.column_dimensions['A'].width = 25

        # Aplicar bordas a todas as células com dados
        for row in range(2, row_idx):
            for col in range(1, len(cabecalhos) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

                # Alinhar valores numéricos à direita
                if col > 1:
                    cell.alignment = Alignment(horizontal='right')

        # Congelar a primeira linha e coluna
        ws.freeze_panes = 'B2'

        # Definir o caminho completo do arquivo
        caminho_arquivo = os.path.join(self.pasta_arquivos, nome_arquivo)

        # Salvar o arquivo
        wb.save(caminho_arquivo)

        print(f"Ordens de produção exportadas com sucesso para: {caminho_arquivo}")
        return caminho_arquivo

    def iniciar_execucao(self):
        """
        Inicia a fase de execução e controle do MRP.

        Returns:
            bool: True se a execução foi iniciada com sucesso, None caso contrário.
        """
        if not hasattr(self, 'estado') or self.estado != "Planejado":
            print("Erro: O planejamento ainda não foi realizado.")
            return None

        # Mudar o estado do MRP para Em Execução
        self.estado = "Em Execução"

        # Inicializar o dicionário de ordens_controle
        self.ordens_controle = {}

        # Copiar as ordens de planejamento para o dicionário de controle
        for material, ordens in self.ordens_planejamento.items():
            self.ordens_controle[material] = {
                'Estoque Atual': self.estoque.get(material, {'em_estoque': 0})['em_estoque'],
                'Status': 'Planejada'  # Status inicial: Planejada
            }

            # Copiar todas as chaves existentes em ordens para ordens_controle
            if 'Produção' in ordens:
                self.ordens_controle[material]['Produção'] = ordens['Produção']

            if 'Aquisição' in ordens:
                self.ordens_controle[material]['Aquisição'] = ordens['Aquisição']

        print("Execução iniciada com sucesso. Estado atual: Em Execução")
        return True

    def editar_ordem(self, material, tipo_ordem, nova_quantidade):
        """
        Edita a quantidade de uma ordem específica.

        Args:
            material (str): Código do material cuja ordem será editada.
            tipo_ordem (str): Tipo de ordem ('Produção' ou 'Aquisição').
            nova_quantidade (int): Nova quantidade para a ordem.

        Returns:
            bool: True se a ordem foi editada com sucesso, False caso contrário.
        """
        if not hasattr(self, 'estado') or self.estado != "Em Execução":
            print("Erro: O MRP não está em execução.")
            return False

        if material not in self.ordens_controle:
            print(f"Erro: Material '{material}' não encontrado nas ordens de controle.")
            return False

        if tipo_ordem not in ['Produção', 'Aquisição']:
            print(f"Erro: Tipo de ordem '{tipo_ordem}' inválido. Use 'Produção' ou 'Aquisição'.")
            return False

        if tipo_ordem not in self.ordens_controle[material]:
            print(f"Erro: Não existe ordem de {tipo_ordem} para o material '{material}'.")
            return False

        # Verificar se a ordem já está pronta
        if self.ordens_controle[material]['Status'] == 'Pronta':
            print(f"Erro: Não é possível editar uma ordem que já está pronta.")
            return False

        # Salvar quantidade anterior para log
        quantidade_anterior = self.ordens_controle[material][tipo_ordem]

        # Atualizar a quantidade
        self.ordens_controle[material][tipo_ordem] = nova_quantidade

        # Atualizar o status para 'Planejada' se estava 'Executada'
        if self.ordens_controle[material]['Status'] == 'Executada':
            self.ordens_controle[material]['Status'] = 'Planejada'
            print(f"Status da ordem para '{material}' atualizado para 'Planejada' devido à edição.")

        print(f"Ordem de {tipo_ordem} para '{material}' editada: {quantidade_anterior} -> {nova_quantidade}")
        return True

    def cancelar_ordem(self, material):
        """
        Cancela uma ordem específica.

        Args:
            material (str): Código do material cuja ordem será cancelada.

        Returns:
            bool: True se a ordem foi cancelada com sucesso, False caso contrário.
        """
        if not hasattr(self, 'estado') or self.estado != "Em Execução":
            print("Erro: O MRP não está em execução.")
            return False

        if material not in self.ordens_controle:
            print(f"Erro: Material '{material}' não encontrado nas ordens de controle.")
            return False

        # Verificar se a ordem já está pronta
        if self.ordens_controle[material]['Status'] == 'Pronta':
            print(f"Erro: Não é possível cancelar uma ordem que já está pronta.")
            return False

        # Salvar informações para log
        tipos_ordem = []
        if 'Produção' in self.ordens_controle[material]:
            tipos_ordem.append('Produção')
        if 'Aquisição' in self.ordens_controle[material]:
            tipos_ordem.append('Aquisição')

        # Cancelar a ordem (zerar quantidades)
        for tipo in tipos_ordem:
            self.ordens_controle[material][tipo] = 0

        # Atualizar o status para 'Planejada'
        self.ordens_controle[material]['Status'] = 'Planejada'

        print(f"Ordem para '{material}' cancelada. Quantidades zeradas e status definido como 'Planejada'.")
        return True

    def atualizar_status_ordem(self, material, novo_status):
        """
        Atualiza o status de uma ordem específica.

        Args:
            material (str): Código do material cuja ordem será atualizada.
            novo_status (str): Novo status da ordem ('Planejada', 'Executada' ou 'Pronta').

        Returns:
            bool: True se o status foi atualizado com sucesso, False caso contrário.
        """
        if not hasattr(self, 'estado') or self.estado != "Em Execução":
            print("Erro: O MRP não está em execução.")
            return False

        if material not in self.ordens_controle:
            print(f"Erro: Material '{material}' não encontrado nas ordens de controle.")
            return False

        # Verificar se o novo status é válido
        status_validos = ['Planejada', 'Executada', 'Pronta']
        if novo_status not in status_validos:
            print(f"Erro: Status '{novo_status}' inválido. Use um dos seguintes: {', '.join(status_validos)}")
            return False

        # Verificar se a ordem tem quantidades válidas
        tem_producao = 'Produção' in self.ordens_controle[material] and self.ordens_controle[material]['Produção'] > 0
        tem_aquisicao = 'Aquisição' in self.ordens_controle[material] and self.ordens_controle[material][
            'Aquisição'] > 0

        if not (tem_producao or tem_aquisicao):
            print(f"Erro: Não é possível atualizar o status de uma ordem cancelada (quantidades zeradas).")
            return False

        # Atualizar o status
        self.ordens_controle[material]['Status'] = novo_status
        print(f"Status da ordem para '{material}' atualizado para '{novo_status}'.")

        # Se o status for 'Pronta', atualizar o estoque
        if novo_status == 'Pronta':
            # Atualizar o estoque com base no tipo de ordem (Produção ou Aquisição)
            if 'Produção' in self.ordens_controle[material] and self.ordens_controle[material]['Produção'] > 0:
                quantidade = self.ordens_controle[material]['Produção']
                self.estoque[material]['em_estoque'] += quantidade
                print(f"Estoque de '{material}' atualizado: +{quantidade} unidades (Produção)")

            if 'Aquisição' in self.ordens_controle[material] and self.ordens_controle[material]['Aquisição'] > 0:
                quantidade = self.ordens_controle[material]['Aquisição']
                self.estoque[material]['em_estoque'] += quantidade
                print(f"Estoque de '{material}' atualizado: +{quantidade} unidades (Aquisição)")

        return True

    def listar_ordens_para_edicao(self):
        """
        Lista todas as ordens que podem ser editadas ou canceladas.

        Returns:
            bool: True se as ordens foram listadas com sucesso, False caso contrário.
        """
        if not hasattr(self, 'ordens_controle') or not self.ordens_controle:
            print("Erro: Não há ordens de controle para listar.")
            return False

        from tabulate import tabulate

        # Preparar os dados para a tabela
        tabela = []
        cabecalho = ["Material", "Estoque Atual", "Produção", "Aquisição", "Status", "Editável"]

        for material, dados in self.ordens_controle.items():
            estoque_atual = dados['Estoque Atual']
            status = dados['Status']

            producao = dados.get('Produção', 0)
            aquisicao = dados.get('Aquisição', 0)

            # Verificar se a ordem pode ser editada
            editavel = "Não" if status == "Pronta" else "Sim"

            linha = [
                material,
                estoque_atual,
                producao if producao > 0 else "",
                aquisicao if aquisicao > 0 else "",
                status,
                editavel
            ]

            tabela.append(linha)

        # Imprimir a tabela
        print(tabulate(tabela, headers=cabecalho, tablefmt="grid"))
        return True

    def listar_ordens_controle(self):
        """
        Lista todas as ordens de controle no console.

        Returns:
            bool: True se as ordens foram listadas com sucesso, False caso contrário.
        """
        if not hasattr(self, 'ordens_controle') or not self.ordens_controle:
            print("Erro: Não há ordens de controle para listar.")
            return False

        from tabulate import tabulate

        # Preparar os dados para a tabela
        tabela = []
        cabecalho = ["Material", "Estoque Atual", "Retirada de Estoque", "Produção", "Aquisição", "Status"]

        for material, dados in self.ordens_controle.items():
            estoque_atual = dados['Estoque Atual']
            status = dados['Status']

            # Calcular a retirada de estoque
            producao = dados.get('Produção', 0)
            aquisicao = dados.get('Aquisição', 0)
            retirada_estoque = min(estoque_atual, producao + aquisicao)

            linha = [
                material,
                estoque_atual,
                retirada_estoque if retirada_estoque > 0 else "",
                producao if producao > 0 else "",
                aquisicao if aquisicao > 0 else "",
                status
            ]

            tabela.append(linha)

        # Imprimir a tabela
        print(tabulate(tabela, headers=cabecalho, tablefmt="grid"))
        return True


    def listar_custos_materiais(self):
        """
        Lista os custos estimados de cada material e o custo total em formato tabular.

        Returns:
            float: Custo total estimado de todos os materiais.
        """
        if not hasattr(self, 'fc_lt_esperados') or not self.fc_lt_esperados:
            print("Erro: Fluxo de caixa e lead times ainda não foram calculados.")
            return None

        from tabulate import tabulate

        # Preparar os dados para a tabela
        tabela = []
        cabecalho = ["Material", "Tipo", "Quantidade", "Custo Estimado (R$)"]
        custo_total = 0

        for material, dados in self.fc_lt_esperados.items():
            if 'Custo' in dados:
                # Determinar o tipo (Produção ou Aquisição)
                tipo = "Produção" if material in self.boms else "Aquisição"

                # Obter a quantidade da ordem
                quantidade = 0
                if material in self.ordens_planejamento:
                    if tipo == "Produção" and 'Produção' in self.ordens_planejamento[material]:
                        quantidade = self.ordens_planejamento[material]['Produção']
                    elif tipo == "Aquisição" and 'Aquisição' in self.ordens_planejamento[material]:
                        quantidade = self.ordens_planejamento[material]['Aquisição']

                # Adicionar à tabela
                custo = dados['Custo']
                tabela.append([material, tipo, quantidade, f"{custo:.2f}"])
                custo_total += custo

        # Adicionar linha de total
        tabela.append(["TOTAL", "", "", f"{custo_total:.2f}"])

        # Imprimir a tabela
        print(tabulate(tabela, headers=cabecalho, tablefmt="grid"))

        return custo_total


    def exportar_custos_materiais(self, nome_arquivo):
        """
        Exporta os custos estimados de cada material e o custo total para uma planilha Excel.

        Args:
            nome_arquivo (str): Nome do arquivo Excel a ser criado.

        Returns:
            str: Caminho completo do arquivo Excel criado.
        """
        if not hasattr(self, 'fc_lt_esperados') or not self.fc_lt_esperados:
            print("Erro: Fluxo de caixa e lead times ainda não foram calculados.")
            return None

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Custos Materiais"

        # Definir cabeçalhos
        cabecalhos = ["Material", "Tipo", "Quantidade", "Custo Estimado (R$)"]

        # Adicionar cabeçalhos
        for col_idx, header in enumerate(cabecalhos, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Preencher os dados
        row_idx = 2
        custo_total = 0

        for material, dados in self.fc_lt_esperados.items():
            if 'Custo' in dados:
                # Determinar o tipo (Produção ou Aquisição)
                tipo = "Produção" if material in self.boms else "Aquisição"

                # Obter a quantidade da ordem
                quantidade = 0
                if material in self.ordens_planejamento:
                    if tipo == "Produção" and 'Produção' in self.ordens_planejamento[material]:
                        quantidade = self.ordens_planejamento[material]['Produção']
                    elif tipo == "Aquisição" and 'Aquisição' in self.ordens_planejamento[material]:
                        quantidade = self.ordens_planejamento[material]['Aquisição']

                # Adicionar à planilha
                ws.cell(row=row_idx, column=1, value=material)
                ws.cell(row=row_idx, column=2, value=tipo)
                ws.cell(row=row_idx, column=3, value=quantidade)

                custo = dados['Custo']
                ws.cell(row=row_idx, column=4, value=custo)
                ws.cell(row=row_idx, column=4).number_format = '#,##0.00'

                custo_total += custo
                row_idx += 1

        # Adicionar linha de total
        ws.cell(row=row_idx, column=1, value="TOTAL")
        ws.cell(row=row_idx, column=4, value=custo_total)
        ws.cell(row=row_idx, column=4).number_format = '#,##0.00'

        # Aplicar formatação
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar estilo aos cabeçalhos
        for col_idx in range(1, len(cabecalhos) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

            # Ajustar largura da coluna
            ws.column_dimensions[cell.column_letter].width = 18

        # Aplicar estilo à linha de total
        for col_idx in range(1, len(cabecalhos) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border

        # Aplicar bordas a todas as células com dados
        for row in range(2, row_idx):
            for col in range(1, len(cabecalhos) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

                # Alinhar valores numéricos à direita
                if col > 2:
                    cell.alignment = Alignment(horizontal='right')

        # Congelar a primeira linha
        ws.freeze_panes = 'A2'

        # Definir o caminho completo do arquivo
        caminho_arquivo = os.path.join(self.pasta_arquivos, nome_arquivo)

        # Salvar o arquivo
        wb.save(caminho_arquivo)

        print(f"Custos de materiais exportados com sucesso para: {caminho_arquivo}")
        return caminho_arquivo

    def atualizar_custos_leadtimes(self, nome_arquivo_cotacoes):
        """
        Atualiza os custos e leadtimes com base na planilha de cotações.

        Args:
            nome_arquivo_cotacoes (str): Nome do arquivo Excel com as cotações.

        Returns:
            tuple: (bool, list) - Sucesso da operação e lista de alertas.
        """
        alertas = []
        caminho_arquivo = os.path.join(self.pasta_arquivos, nome_arquivo_cotacoes)

        try:
            # Carregar a planilha de cotações
            wb = openpyxl.load_workbook(caminho_arquivo)
            ws = wb.active

            # Obter os cabeçalhos
            headers = [cell.value for cell in ws[1]]

            # Processar cada linha da planilha
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Pular linhas vazias
                    continue

                material = row[0]

                if material in self.estoque:
                    # Mapear os valores da linha para um dicionário
                    row_data = dict(zip(headers, row))

                    # Atualizar custo unitário
                    if 'Custo Unitario' in row_data and row_data['Custo Unitario'] is not None:
                        custo_unitario = row_data['Custo Unitario']
                        if custo_unitario != self.estoque[material]['custo_medio_unitario']:
                            alerta = f"Custo de {material} alterado: {self.estoque[material]['custo_medio_unitario']} -> {custo_unitario}"
                            alertas.append(alerta)
                            self.estoque[material]['custo_medio_unitario'] = custo_unitario

                    # Atualizar imposto unitário
                    if 'Imposto Unitario' in row_data and row_data['Imposto Unitario'] is not None:
                        imposto_unitario = row_data['Imposto Unitario']
                        if imposto_unitario != self.estoque[material]['imposto_medio_unitario']:
                            alertas.append(
                                f"Imposto de {material} alterado: {self.estoque[material]['imposto_medio_unitario']} -> {imposto_unitario}")
                            self.estoque[material]['imposto_medio_unitario'] = imposto_unitario

                    # Atualizar frete por lote
                    if 'Frete Lote' in row_data and row_data['Frete Lote'] is not None:
                        frete_lote = row_data['Frete Lote']
                        if frete_lote != self.estoque[material]['frete_medio_lote']:
                            alertas.append(
                                f"Frete de {material} alterado: {self.estoque[material]['frete_medio_lote']} -> {frete_lote}")
                            self.estoque[material]['frete_medio_lote'] = frete_lote

                    # Atualizar leadtime
                    if 'Lead Time' in row_data and row_data['Lead Time'] is not None:
                        lead_time = row_data['Lead Time']
                        if lead_time > self.estoque[material]['leadtime_medio_lote']:
                            alertas.append(
                                f"Leadtime de {material} aumentou: {self.estoque[material]['leadtime_medio_lote']} -> {lead_time}")
                            self.estoque[material]['leadtime_medio_lote'] = lead_time
                            alertas.append("Necessidade de replanejar devido a aumento no leadtime.")
                        elif lead_time != self.estoque[material]['leadtime_medio_lote']:
                            alertas.append(
                                f"Leadtime de {material} alterado: {self.estoque[material]['leadtime_medio_lote']} -> {lead_time}")
                            self.estoque[material]['leadtime_medio_lote'] = lead_time

            if alertas:
                print("Alertas durante a atualização:")
                for alerta in alertas:
                    print(f"- {alerta}")

            return True, alertas

        except Exception as e:
            print(f"Erro ao atualizar custos e leadtimes: {str(e)}")
            return False, []

    def recuperar_planejamento(self, nome_arquivo):
        """
        Recupera o planejamento de uma planilha Excel e o carrega no dicionário de planejamento.

        Args:
            nome_arquivo (str): Nome do arquivo Excel contendo o planejamento.

        Returns:
            bool: True se o planejamento foi recuperado com sucesso, False caso contrário.
        """
        import pandas as pd
        from datetime import datetime

        caminho_arquivo = os.path.join(self.pasta_arquivos, nome_arquivo)

        try:
            # Lê a planilha Excel
            df = pd.read_excel(caminho_arquivo)

            # Reinicializa o dicionário de planejamento
            self.planejamento = {}

            # Itera sobre as linhas do DataFrame
            for _, row in df.iterrows():
                material = row['Material']
                self.planejamento[material] = {'Estoque Atual': row['Estoque Atual']}

                # Adiciona as datas e quantidades ao dicionário
                for col in df.columns:
                    if col not in ['Material', 'Estoque Atual']:
                        try:
                            data = datetime.strptime(col, '%Y-%m-%d')
                            quantidade = row[col]
                            if pd.notna(quantidade):
                                self.planejamento[material][col] = quantidade
                        except ValueError:
                            # Ignora colunas que não são datas
                            pass

            print(f"Planejamento recuperado com sucesso de: {caminho_arquivo}")
            return True

        except Exception as e:
            print(f"Erro ao recuperar o planejamento: {str(e)}")
            return False

    def planejar_producao(self, demanda):
        """
        Realiza o planejamento da produção com base na demanda e nos dados inicializados.

        Args:
            demanda (dict): Dicionário contendo os produtos finais e suas respectivas quantidades demandadas.
                            Exemplo: {"ETF": 10, "ETI": 15}
            """
        if self.estado != "Inicializado":
            print("Erro: Os dados ainda não foram inicializados.")
            return
        # Calcula as quantidades de produção e aquisição
        self.calcular_quantidades_producao_aquisicao(demanda)
        # Calcula leadtimes
        self.calcular_fc_lt_esperados()
        # Monta o quadro de planejamento
        self.montar_quadro_planejamento()
        self.estado = "Planejado"

    def executar_controle(self, cotacoes):
        """
        Executa e controla as ordens de produção e aquisição.

        Args:
            cotacoes (dict): Dicionário contendo os valores atualizados dos materiais.
        """
        if self.estado != "Planejado":
            print("Erro: O planejamento ainda não foi realizado.")
            return
        pass

    def analisar_resultados(self):
        """
        Analisa os resultados do ciclo de produção.
        """
        if self.estado != "Encerrado":
            print("Erro: O ciclo de produção ainda não foi encerrado.")
            return
        pass
import os
import pandas as pd
from datetime import datetime


class CRP:
    """
    Classe que implementa o Capacity Requirements Planning (CRP).

    O CRP é responsável por ajustar o planejamento do MRP considerando
    as restrições de capacidade dos recursos produtivos.
    """

    def __init__(self, pasta_arquivos):
        """
        Inicializa o CRP.

        Args:
            pasta_arquivos (str): Caminho para a pasta onde os arquivos serão lidos/salvos.
        """
        self.pasta_arquivos = pasta_arquivos
        self.estado = "Não Inicializado"
        self.planejamento_mrp = {}
        self.datas_entrega = []
        self.produtos = []

    def carregar_planejamento_mrp(self, nome_arquivo):
        """
        Carrega o planejamento do MRP a partir de uma planilha Excel.

        Args:
            nome_arquivo (str): Nome do arquivo Excel contendo o planejamento do MRP.

        Returns:
            bool: True se o planejamento foi carregado com sucesso, False caso contrário.
        """
        try:
            # Constrói o caminho completo do arquivo
            caminho_arquivo = os.path.join(self.pasta_arquivos, nome_arquivo)

            # Lê a planilha Excel
            df = pd.read_excel(caminho_arquivo)

            # Inicializa o dicionário de planejamento
            self.planejamento_mrp = {}

            # Extrai as datas de entrega (todas as colunas exceto 'Material' e 'Estoque Atual')
            self.datas_entrega = [col for col in df.columns if col not in ['Material', 'Estoque Atual']]

            # Extrai os produtos (todos os valores da coluna 'Material')
            self.produtos = df['Material'].tolist()

            # Itera sobre as linhas do DataFrame
            for _, row in df.iterrows():
                material = row['Material']
                self.planejamento_mrp[material] = {'Estoque Atual': row['Estoque Atual']}

                # Adiciona as datas e quantidades ao dicionário
                for data in self.datas_entrega:
                    quantidade = row[data]
                    if pd.notna(quantidade) and quantidade > 0:
                        self.planejamento_mrp[material][data] = quantidade

            # Atualiza o estado para "Inicializado"
            self.estado = "Inicializado"

            print(f"Planejamento do MRP carregado com sucesso de: {caminho_arquivo}")
            return True

        except Exception as e:
            print(f"Erro ao carregar o planejamento do MRP: {str(e)}")
            return False

    def carregar_demanda_recursos(self, nome_arquivo):
        """
        Carrega a planilha de demanda por recursos.

        Args:
            nome_arquivo (str): Nome do arquivo Excel contendo a demanda por recursos.

        Returns:
            bool: True se a demanda foi carregada com sucesso, False caso contrário.
        """
        try:
            caminho_arquivo = os.path.join(self.pasta_arquivos, nome_arquivo)
            df = pd.read_excel(caminho_arquivo)

            self.demanda_recursos = {}
            for _, row in df.iterrows():
                produto = row['Produto']
                self.demanda_recursos[produto] = {}
                for col in df.columns:
                    if col != 'Produto' and pd.notna(row[col]):
                        self.demanda_recursos[produto][col] = row[col]

            print(f"Demanda por recursos carregada com sucesso de: {caminho_arquivo}")
            return True
        except Exception as e:
            print(f"Erro ao carregar a demanda por recursos: {str(e)}")
            return False

    def calcular_demanda_por_operacao(self):
        """
        Calcula a demanda por operação com base no planejamento do MRP e na demanda por recursos.

        Returns:
            dict: Demanda por operação, organizada por produto e operação.
        """
        if not hasattr(self, 'demanda_recursos'):
            print("Erro: Demanda por recursos não foi carregada.")
            return None

        demanda_por_operacao = {}

        for produto, planejamento in self.planejamento_mrp.items():
            if produto not in self.demanda_recursos:
                continue

            total_unidades = sum(
                quantidade for data, quantidade in planejamento.items()
                if data != "Estoque Atual"
            )

            for operacao, minutos_por_unidade in self.demanda_recursos[produto].items():
                if operacao not in demanda_por_operacao:
                    demanda_por_operacao[operacao] = {}

                demanda_por_operacao[operacao][produto] = total_unidades * minutos_por_unidade

        return demanda_por_operacao

    def carregar_capacidade_recursos(self, nome_arquivo):
        """
        Carrega a planilha de capacidade de produção nominal por recurso e operação.

        Args:
            nome_arquivo (str): Nome do arquivo Excel contendo a capacidade dos recursos.

        Returns:
            bool: True se a capacidade foi carregada com sucesso, False caso contrário.
        """
        try:
            caminho_arquivo = os.path.join(self.pasta_arquivos, nome_arquivo)
            df = pd.read_excel(caminho_arquivo)

            # Inicializa o dicionário de capacidade de recursos
            self.capacidade_recursos = {}

            # A primeira coluna deve conter os recursos
            recursos = df.iloc[:, 0].tolist()

            # As demais colunas são as operações
            operacoes = df.columns[1:].tolist()

            # Preenche o dicionário de capacidade
            for i, recurso in enumerate(recursos):
                self.capacidade_recursos[recurso] = {}
                for operacao in operacoes:
                    self.capacidade_recursos[recurso][operacao] = df.iloc[i][operacao]

            print(f"Capacidade de recursos carregada com sucesso de: {caminho_arquivo}")
            return True

        except Exception as e:
            print(f"Erro ao carregar a capacidade de recursos: {str(e)}")
            return False


    def carregar_excecoes_capacidade(self, nome_arquivo):
        """
        Carrega a planilha de exceções de capacidade, que contém reduções na capacidade
        nominal dos recursos em datas específicas.

        Args:
            nome_arquivo (str): Nome do arquivo Excel contendo as exceções de capacidade.

        Returns:
            bool: True se as exceções foram carregadas com sucesso, False caso contrário.
        """
        try:
            caminho_arquivo = os.path.join(self.pasta_arquivos, nome_arquivo)
            df = pd.read_excel(caminho_arquivo, header=None, parse_dates=False)

            self.excecoes_capacidade = {}
            recurso_atual = None
            operacoes = []

            for _, row in df.iterrows():
                # Se a primeira célula contém um recurso
                if pd.notna(row.iloc[0]) and isinstance(row.iloc[0], str) and not row.iloc[0].startswith('20'):
                    recurso_atual = row.iloc[0]
                    operacoes = [op for op in row.iloc[1:].tolist() if pd.notna(op)]
                    self.excecoes_capacidade[recurso_atual] = {}
                    for op in operacoes:
                        self.excecoes_capacidade[recurso_atual][op] = {}
                # Se a primeira célula contém uma data
                elif recurso_atual and pd.notna(row.iloc[0]):
                    # Converter a data para string no formato YYYY-MM-DD
                    if isinstance(row.iloc[0], pd.Timestamp):
                        data = row.iloc[0].strftime('%Y-%m-%d')
                    else:
                        try:
                            data = pd.to_datetime(row.iloc[0]).strftime('%Y-%m-%d')
                        except:
                            data = str(row.iloc[0])

                    # Processar cada operação
                    for i, operacao in enumerate(operacoes):
                        if i + 1 < len(row) and pd.notna(row.iloc[i + 1]):
                            self.excecoes_capacidade[recurso_atual][operacao][data] = row.iloc[i + 1]

            print(f"Exceções de capacidade carregadas com sucesso de: {caminho_arquivo}")
            return True

        except Exception as e:
            print(f"Erro ao carregar as exceções de capacidade: {str(e)}")
            return False

    def criar_planilha_crp(self, nome_arquivo, data_planejamento, numero_dias):
        """
        Cria uma planilha para o CRP com base nos dados de capacidade e demanda calculados anteriormente.

        Args:
            nome_arquivo (str): Nome do arquivo Excel a ser criado.
            data_planejamento (str): Data inicial do planejamento (formato YYYY-MM-DD).
            numero_dias (int): Número de dias para o planejamento.

        Returns:
            str: Caminho completo do arquivo Excel criado.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from datetime import datetime, timedelta

        # Verificar se todos os dados necessários foram carregados
        if not hasattr(self, 'capacidade_recursos') or not hasattr(self, 'demanda_recursos'):
            print("Erro: Capacidade de recursos ou demanda por recursos não foram carregados.")
            return None

        # Identificar produtos finais (aqueles que estão na demanda_recursos)
        produtos_finais = list(self.demanda_recursos.keys())

        # Criar um novo workbook
        wb = openpyxl.Workbook()

        # Criar aba de demanda total
        ws_demanda = wb.active
        ws_demanda.title = "Demanda Total"

        # Cabeçalhos da aba de demanda
        cabecalhos_demanda = ["Produto", "Demanda Total", "Alocado", "Pendente"]

        for col_num, cabecalho in enumerate(cabecalhos_demanda, 1):
            cell = ws_demanda.cell(row=1, column=col_num, value=cabecalho)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        # Preencher dados de produtos finais e demanda total
        linha_atual = 2
        for produto in produtos_finais:
            if produto in self.planejamento_mrp:
                planejamento = self.planejamento_mrp[produto]
                # Calcular demanda total
                demanda_total = sum(quantidade for data, quantidade in planejamento.items()
                                    if data != "Estoque Atual")

                if demanda_total > 0:
                    ws_demanda.cell(row=linha_atual, column=1, value=produto)
                    ws_demanda.cell(row=linha_atual, column=2, value=demanda_total)

                    # Fórmula para calcular o total alocado (será preenchida depois)
                    ws_demanda.cell(row=linha_atual, column=3, value=0)

                    # Fórmula para calcular o pendente
                    ws_demanda.cell(row=linha_atual, column=4, value=f"=B{linha_atual}-C{linha_atual}")
                    linha_atual += 1

        # Ajustar largura das colunas
        for col_num in range(1, len(cabecalhos_demanda) + 1):
            ws_demanda.column_dimensions[get_column_letter(col_num)].width = 20

        # Adicionar bordas
        borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws_demanda[f"A1:D{linha_atual - 1}"]:
            for cell in row:
                cell.border = borda_fina

        # Criar planilhas de alocação para cada dia
        for dia in range(numero_dias):
            data_atual = datetime.strptime(data_planejamento, '%Y-%m-%d') + timedelta(days=dia)
            data_str = data_atual.strftime('%Y-%m-%d')
            ws = wb.create_sheet(title=f"Alocação {data_str}")

            # Cabeçalhos da planilha de alocação
            cabecalhos = ["Produto", "Demanda Pendente", "Quantidade Alocada"]

            for col_num, cabecalho in enumerate(cabecalhos, 1):
                cell = ws.cell(row=1, column=col_num, value=cabecalho)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

            # Preencher dados de produtos finais e demanda
            linha_atual = 2
            for i, produto in enumerate(produtos_finais):
                if produto in self.planejamento_mrp:
                    produto_idx = i + 2  # +2 porque a linha 1 é cabeçalho
                    ws.cell(row=linha_atual, column=1, value=produto)
                    ws.cell(row=linha_atual, column=2, value=f"='Demanda Total'!D{produto_idx}")
                    ws.cell(row=linha_atual, column=3, value=0)
                    linha_atual += 1

            # Adicionar validação de dados para "Quantidade Alocada"
            dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1=0)
            dv.error = "A quantidade alocada deve ser um número inteiro não negativo."
            dv.errorTitle = "Entrada inválida"
            ws.add_data_validation(dv)
            dv.add(f"C2:C{linha_atual - 1}")

            # Ajustar largura das colunas
            for col_num in range(1, len(cabecalhos) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 20

            # Adicionar bordas
            for row in ws[f"A1:C{linha_atual - 1}"]:
                for cell in row:
                    cell.border = borda_fina

            # Adicionar tabela de alocação de recursos para este dia
            linha_recursos = linha_atual + 2
            ws.cell(row=linha_recursos, column=1, value="Alocação de Recursos")
            ws.cell(row=linha_recursos, column=1).font = Font(bold=True)

            # Cabeçalhos da tabela de recursos
            cabecalhos_recursos = ["Recurso", "Operação", "Capacidade Nominal", "Exceção",
                                   "Capacidade Disponível", "Consumo", "% Utilização"]

            for col_num, cabecalho in enumerate(cabecalhos_recursos, 1):
                cell = ws.cell(row=linha_recursos + 1, column=col_num, value=cabecalho)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

            # Preencher dados de recursos
            linha_atual_recursos = linha_recursos + 2
            for recurso, operacoes in self.capacidade_recursos.items():
                for operacao, capacidade_nominal in operacoes.items():
                    if capacidade_nominal > 0:  # Só incluir operações que o recurso pode realizar
                        # Verificar se há exceção para este recurso/operação nesta data
                        excecao = 0
                        if hasattr(self, 'excecoes_capacidade') and recurso in self.excecoes_capacidade:
                            if operacao in self.excecoes_capacidade[recurso]:
                                if data_str in self.excecoes_capacidade[recurso][operacao]:
                                    excecao = self.excecoes_capacidade[recurso][operacao][data_str]

                        # Calcular capacidade disponível
                        capacidade_disponivel = max(0, capacidade_nominal - excecao)

                        # Fórmula para calcular o consumo baseado nas quantidades alocadas
                        formula_consumo = "="
                        primeiro = True

                        for i, produto in enumerate(produtos_finais):
                            if produto in self.demanda_recursos and operacao in self.demanda_recursos[produto]:
                                minutos_por_unidade = self.demanda_recursos[produto][operacao]
                                # Encontrar a linha do produto na planilha de alocação
                                produto_linha = i + 2  # +2 porque a linha 1 é cabeçalho
                                if not primeiro:
                                    formula_consumo += "+"
                                formula_consumo += f"C{produto_linha}*{minutos_por_unidade}"
                                primeiro = False

                        if formula_consumo == "=":
                            formula_consumo = "=0"

                        ws.cell(row=linha_atual_recursos, column=1, value=recurso)
                        ws.cell(row=linha_atual_recursos, column=2, value=operacao)
                        ws.cell(row=linha_atual_recursos, column=3, value=capacidade_nominal)
                        ws.cell(row=linha_atual_recursos, column=4, value=excecao)
                        ws.cell(row=linha_atual_recursos, column=5, value=capacidade_disponivel)
                        ws.cell(row=linha_atual_recursos, column=6, value=formula_consumo)
                        ws.cell(row=linha_atual_recursos, column=7,
                                value=f"=F{linha_atual_recursos}/E{linha_atual_recursos}")
                        ws.cell(row=linha_atual_recursos, column=7).number_format = "0.00%"

                        linha_atual_recursos += 1

            # Ajustar largura das colunas
            for col_num in range(1, len(cabecalhos_recursos) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 20

            # Adicionar bordas
            for row in ws[f"A{linha_recursos + 1}:G{linha_atual_recursos - 1}"]:
                for cell in row:
                    cell.border = borda_fina

            # Adicionar formatação condicional para % Utilização
            from openpyxl.formatting.rule import ColorScaleRule
            color_scale = ColorScaleRule(start_type='num', start_value=0, start_color='00FF00',
                                         mid_type='num', mid_value=0.7, mid_color='FFFF00',
                                         end_type='num', end_value=1, end_color='FF0000')
            ws.conditional_formatting.add(f"G{linha_recursos + 2}:G{linha_atual_recursos - 1}", color_scale)

        # Agora que todas as abas foram criadas, adicionar fórmulas para atualizar a aba de Demanda Total
        for i, produto in enumerate(produtos_finais):
            if produto in self.planejamento_mrp:
                produto_idx = i + 2  # +2 porque a linha 1 é cabeçalho

                # Construir fórmula para somar todas as quantidades alocadas nas abas diárias
                formula_alocado = "="
                primeiro = True

                for dia in range(numero_dias):
                    data_atual = datetime.strptime(data_planejamento, '%Y-%m-%d') + timedelta(days=dia)
                    data_str = data_atual.strftime('%Y-%m-%d')

                    if not primeiro:
                        formula_alocado += "+"
                    formula_alocado += f"'Alocação {data_str}'!C{produto_idx}"
                    primeiro = False

                # Atualizar a célula com a fórmula
                ws_demanda.cell(row=produto_idx, column=3, value=formula_alocado)

        # Remover a planilha padrão criada pelo openpyxl se existir
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Definir o caminho completo do arquivo
        caminho_arquivo = os.path.join(self.pasta_arquivos, nome_arquivo)

        # Salvar o arquivo
        wb.save(caminho_arquivo)

        print(f"Planilha CRP criada com sucesso: {caminho_arquivo}")
        return caminho_arquivo

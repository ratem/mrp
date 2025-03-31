'''
import os
from datetime import datetime
from mrp import MRP

if __name__ == '__main__':
    print("=" * 80)
    print("SIMULAÇÃO DO SISTEMA MRP - MATERIAL REQUIREMENTS PLANNING")
    print("=" * 80)

    # Definir o diretório de trabalho e criar uma pasta para o ciclo atual
    diretorio_base = os.getcwd()
    data_ciclo = datetime.now().strftime("%Y%m%d_%H%M")
    pasta_ciclo = os.path.join(diretorio_base, f"ciclo_{data_ciclo}")

    if not os.path.exists(pasta_ciclo):
        os.makedirs(pasta_ciclo)

    # Criar instância do MRP
    mrp = MRP(pasta_ciclo)

    # Copiar arquivos necessários para a pasta do ciclo
    for arquivo in ["Estoque.xlsx", "ETI_BOM.xlsx", "ETF_BOM.xlsx", "Cotacoes.xlsx"]:
        if os.path.exists(os.path.join(diretorio_base, arquivo)):
            import shutil

            shutil.copy(os.path.join(diretorio_base, arquivo), os.path.join(pasta_ciclo, arquivo))

    print("\n" + "=" * 80)
    print("FASE 1: INICIALIZAÇÃO")
    print("=" * 80)
    print("Carregando dados de estoque e BOMs...")
    mrp.inicializar_dados()
    print("Dados inicializados com sucesso.")

    # Definir a demanda
    demanda = {"ETI": 100, "ETF": 150}
    print(f"\nDemanda definida: {demanda}")

    print("\n" + "=" * 80)
    print("FASE 2: PLANEJAMENTO INICIAL")
    print("=" * 80)
    print("Realizando planejamento inicial...")

    # Realizar o planejamento
    mrp.planejar_producao(demanda)

    # Listar o quadro de planejamento
    mrp.imprimir_quadro_planejamento()

    # Exportar o quadro de planejamento inicial
    arquivo_planejamento_inicial = "planejamento_inicial.xlsx"
    mrp.exportar_quadro_planejamento(arquivo_planejamento_inicial)
    print(f"Quadro de planejamento inicial exportado para: {arquivo_planejamento_inicial}")

    # Exportar as ordens iniciais
    arquivo_ordens_inicial = "ordens_inicial.xlsx"
    mrp.exportar_ordens_producao(arquivo_ordens_inicial)
    print(f"Ordens de produção iniciais exportadas para: {arquivo_ordens_inicial}")

    # Mostrar custos estimados
    print("\nCustos estimados no planejamento inicial:")
    custo_total_inicial = mrp.listar_custos_materiais()
    mrp.exportar_custos_materiais("custos_inicial.xlsx")

    print("\n" + "=" * 80)
    print("FASE 3: EXECUÇÃO E CONTROLE")
    print("=" * 80)
    print("Iniciando execução do planejamento...")

    # Iniciar a execução
    mrp.iniciar_execucao()

    # Listar ordens de controle
    print("\nOrdens de controle no início da execução:")
    mrp.listar_ordens_controle()

    # Atualizar o status de algumas ordens para "Executada"
    print("\nAtualizando status de algumas ordens para 'Executada'...")
    materiais_em_execucao = ["JOKER", "DAQ", "ADS1115"]
    for material in materiais_em_execucao:
        mrp.atualizar_status_ordem(material, "Executada")

    print("\nOrdens de controle após atualização de status:")
    mrp.listar_ordens_controle()

    # Atualizar custos e leadtimes com base em novas cotações
    print("\nAtualizando custos e leadtimes com base em novas cotações...")
    sucesso, alertas = mrp.atualizar_custos_leadtimes("Cotacoes.xlsx")

    # Verificar se há necessidade de replanejamento
    necessidade_replanejar = any("Necessidade de replanejar" in alerta for alerta in alertas)

    if necessidade_replanejar:
        print("\n" + "=" * 80)
        print("REPLANEJAMENTO NECESSÁRIO")
        print("=" * 80)
        print("Detectado aumento significativo em leadtimes. Realizando replanejamento...")

        # Salvar o planejamento atual
        mrp.exportar_quadro_planejamento("planejamento_pre_atualizacao.xlsx")

        # Realizar novo planejamento com os valores atualizados
        mrp.planejar_producao(demanda)

        # Exportar o novo quadro de planejamento
        arquivo_planejamento_atualizado = "planejamento_atualizado.xlsx"
        mrp.exportar_quadro_planejamento(arquivo_planejamento_atualizado)
        print(f"Novo quadro de planejamento exportado para: {arquivo_planejamento_atualizado}")

        # Exportar as novas ordens
        arquivo_ordens_atualizado = "ordens_atualizado.xlsx"
        mrp.exportar_ordens_producao(arquivo_ordens_atualizado)
        print(f"Novas ordens de produção exportadas para: {arquivo_ordens_atualizado}")

        # Mostrar novos custos estimados
        print("\nCustos estimados após replanejamento:")
        custo_total_atualizado = mrp.listar_custos_materiais()
        mrp.exportar_custos_materiais("custos_atualizado.xlsx")

        # Calcular a diferença de custo
        diferenca_custo = custo_total_atualizado - custo_total_inicial
        print(f"\nDiferença de custo após replanejamento: R$ {diferenca_custo:.2f}")

        # Reiniciar a execução com o novo planejamento
        print("\nReiniciando execução com o novo planejamento...")
        mrp.iniciar_execucao()

    # Continuar a execução
    print("\nContinuando a execução...")

    # Atualizar o status de algumas ordens para "Executada"
    materiais_em_execucao = ["JOKER", "DAQ", "ADS1115", "Invólucro", "Antena LoRa"]
    for material in materiais_em_execucao:
        mrp.atualizar_status_ordem(material, "Executada")

    # Finalizar algumas ordens (status "Pronta")
    print("\nFinalizando algumas ordens (status 'Pronta')...")
    materiais_prontos = ["JOKER", "DAQ"]
    for material in materiais_prontos:
        mrp.atualizar_status_ordem(material, "Pronta")

    # Verificar a atualização do estoque
    print("\nOrdens de controle após finalização de algumas ordens:")
    mrp.listar_ordens_controle()

    # Editar uma ordem
    print("\nEditando a ordem de 'ADS1115'...")
    mrp.editar_ordem("ADS1115", "Aquisição", 450)

    # Cancelar uma ordem
    print("\nCancelando a ordem de 'Invólucro'...")
    mrp.cancelar_ordem("Invólucro")

    # Listar ordens após edições
    print("\nOrdens de controle após edições:")
    mrp.listar_ordens_controle()

    # Exportar o estado final das ordens
    arquivo_ordens_final = "ordens_final.xlsx"
    mrp.exportar_ordens_producao(arquivo_ordens_final)
    print(f"Estado final das ordens exportado para: {arquivo_ordens_final}")

    print("\n" + "=" * 80)
    print("FASE 4: ANÁLISE")
    print("=" * 80)
    print("Análise do ciclo de produção:")
    print("- Foram detectadas alterações significativas em custos e leadtimes")
    print("- Um replanejamento foi necessário devido ao aumento de leadtimes")
    print(f"- A diferença de custo após o replanejamento foi de R$ {diferenca_custo:.2f}")
    print("- Algumas ordens foram concluídas com sucesso")
    print("- Algumas ordens foram editadas ou canceladas durante a execução")
    print("\nCiclo de MRP concluído com sucesso!")
'''

import os
from datetime import datetime
from mrp import MRP
from crp import CRP

if __name__ == '__main__':
    print("=" * 80)
    print("SIMULAÇÃO INTEGRADA MRP-CRP")
    print("=" * 80)

    # Definir o diretório de trabalho e criar pastas para os ciclos
    diretorio_base = os.getcwd()
    data_ciclo = datetime.now().strftime("%Y%m%d_%H%M")
    pasta_ciclo_mrp = os.path.join(diretorio_base, f"ciclo_mrp_{data_ciclo}")
    pasta_ciclo_crp = os.path.join(diretorio_base, f"ciclo_crp_{data_ciclo}")

    # Criar diretórios se não existirem
    if not os.path.exists(pasta_ciclo_mrp):
        os.makedirs(pasta_ciclo_mrp)
    if not os.path.exists(pasta_ciclo_crp):
        os.makedirs(pasta_ciclo_crp)

    # Criar instâncias do MRP e CRP
    mrp = MRP(pasta_ciclo_mrp)
    crp = CRP(pasta_ciclo_crp)

    # Copiar arquivos necessários para as pastas dos ciclos
    arquivos_mrp = ["Estoque.xlsx", "ETI_BOM.xlsx", "ETF_BOM.xlsx", "Cotacoes.xlsx"]
    for arquivo in arquivos_mrp:
        if os.path.exists(os.path.join(diretorio_base, arquivo)):
            import shutil

            shutil.copy(os.path.join(diretorio_base, arquivo), os.path.join(pasta_ciclo_mrp, arquivo))

    print("\n" + "=" * 80)
    print("FASE 1: CICLO MRP")
    print("=" * 80)

    # Inicializar dados do MRP
    print("Inicializando dados do MRP...")
    mrp.inicializar_dados()

    # Definir a demanda
    demanda = {"ETI": 100, "ETF": 155}
    print(f"Demanda definida: {demanda}")

    # Realizar o planejamento MRP
    print("\nRealizando planejamento MRP...")
    mrp.planejar_producao(demanda)

    # Exibir resultados do MRP em formato tabular
    print("\nQuadro de planejamento MRP:")
    mrp.imprimir_quadro_planejamento()

    # Iniciar execução para poder listar ordens
    mrp.iniciar_execucao()

    print("\nOrdens de produção e aquisição:")
    mrp.listar_ordens_controle()

    print("\nCustos estimados:")
    mrp.listar_custos_materiais()

    # Exportar o quadro de planejamento
    arquivo_planejamento = "planejamento_atualizado.xlsx"
    caminho_planejamento = mrp.exportar_quadro_planejamento(arquivo_planejamento)
    print(f"\nQuadro de planejamento exportado para: {arquivo_planejamento}")

    # Copiar o planejamento para a pasta do CRP
    import shutil

    shutil.copy(caminho_planejamento, os.path.join(pasta_ciclo_crp, arquivo_planejamento))

    print("\n" + "=" * 80)
    print("FASE 2: CICLO CRP")
    print("=" * 80)

    # Criar arquivos necessários para o CRP
    print("Criando arquivos de configuração para o CRP...")

    # Criar arquivo de demanda por recursos
    from openpyxl import Workbook

    # Demanda por recursos
    wb_demanda = Workbook()
    ws_demanda = wb_demanda.active
    ws_demanda.append(["Produto", "OP1", "OP2"])
    ws_demanda.append(["ETI", 20, 40])
    ws_demanda.append(["ETF", 30, 30])
    caminho_demanda = os.path.join(pasta_ciclo_crp, "demanda_recursos.xlsx")
    wb_demanda.save(caminho_demanda)
    print("Arquivo de demanda por recursos criado.")

    # Capacidade de recursos
    wb_capacidade = Workbook()
    ws_capacidade = wb_capacidade.active
    ws_capacidade.append(["Recurso", "OP1", "OP2", "OP3"])
    ws_capacidade.append(["RE1", 480, 0, 480])
    ws_capacidade.append(["RE2", 0, 480, 480])
    caminho_capacidade = os.path.join(pasta_ciclo_crp, "capacidade_recursos.xlsx")
    wb_capacidade.save(caminho_capacidade)
    print("Arquivo de capacidade de recursos criado.")

    # Exceções de capacidade
    wb_excecoes = Workbook()
    ws_excecoes = wb_excecoes.active

    # Adicionar dados para RE1
    ws_excecoes.append(["RE1", "OP1", "OP3"])
    ws_excecoes.append(["2025-04-01", 120, 120])  # Redução significativa no primeiro dia
    ws_excecoes.append(["2025-04-02", 60, 60])  # Redução média no segundo dia
    ws_excecoes.append(["2025-04-03", 30, 30])  # Redução pequena no terceiro dia

    # Adicionar dados para RE2
    ws_excecoes.append(["RE2", "OP2", "OP3"])
    ws_excecoes.append(["2025-04-01", 30, 30])  # Redução pequena no primeiro dia
    ws_excecoes.append(["2025-04-02", 180, 180])  # Redução grande no segundo dia
    ws_excecoes.append(["2025-04-03", 90, 90])  # Redução média no terceiro dia

    caminho_excecoes = os.path.join(pasta_ciclo_crp, "excecoes_capacidade.xlsx")
    wb_excecoes.save(caminho_excecoes)
    print("Arquivo de exceções de capacidade criado com reduções significativas nos primeiros 3 dias.")

    # Inicializar o CRP com os dados do MRP
    print("\nInicializando o CRP com o planejamento do MRP...")
    crp.carregar_planejamento_mrp(arquivo_planejamento)

    # Carregar dados de demanda por recursos
    print("Carregando demanda por recursos...")
    crp.carregar_demanda_recursos("demanda_recursos.xlsx")

    # Carregar dados de capacidade de recursos
    print("Carregando capacidade de recursos...")
    crp.carregar_capacidade_recursos("capacidade_recursos.xlsx")

    # Carregar exceções de capacidade
    print("Carregando exceções de capacidade...")
    crp.carregar_excecoes_capacidade("excecoes_capacidade.xlsx")

    # Calcular demanda por operação
    print("\nCalculando demanda por operação...")
    demanda_por_operacao = crp.calcular_demanda_por_operacao()

    # Exibir demanda por operação
    print("\nDemanda por operação:")
    for operacao, produtos in demanda_por_operacao.items():
        print(f"  {operacao}:")
        for produto, minutos in produtos.items():
            print(f"    {produto}: {minutos} minutos")

    # Criar planilha CRP para planejamento interativo
    print("\nCriando planilha CRP para planejamento interativo...")
    data_atual = datetime.now().strftime("%Y-%m-%d")
    arquivo_crp = "crp_planejamento.xlsx"
    dias_planejamento = 5
    caminho_crp = crp.criar_planilha_crp(arquivo_crp, data_atual, dias_planejamento)

    print(f"\nPlanilha CRP criada em: {arquivo_crp}")
    print("\nInstruções para uso da planilha CRP:")
    print("1. Abra a planilha CRP gerada")
    print(f"2. Nas {dias_planejamento} abas de alocação diária, insira as quantidades de produtos a serem produzidos")
    print("3. Observe na aba 'Capacidade de Recursos' o impacto nas capacidades dos recursos")
    print("4. Ajuste as quantidades para otimizar a utilização dos recursos")
    print("5. Salve a planilha para manter o registro do planejamento de capacidade")

    print("\n" + "=" * 80)
    print("SIMULAÇÃO INTEGRADA MRP-CRP CONCLUÍDA")
    print("=" * 80)
